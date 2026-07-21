"""Portfolio analysis from the Multileg Orders (MLOB) file.

The MLOB has no algo column, so the first step derives one: each
(User ID, Server) is matched against the Compiled User MTM exactly like the
trade value rows; users with no MTM entry inherit their server's algo (same
fallback as the strikes table).

Only COMPLETE rows count. Per portfolio (and per user):

    sell value = sum(Avg Price x Filled Quantity)   over SELL rows
    buy value  = sum(Avg Price x Filled Quantity)   over BUY rows
    PnL        = sell value - buy value

The default report covers the "QS" portfolios — every row whose Portfolio
Name contains "QS" — but portfolio_report() takes any pattern, and the HTML
report embeds the per-(algo, server, user, portfolio) groups so any other
portfolio can be analysed instantly in the browser.
"""

from collections import Counter

import pandas as pd

from tradevalue import INDIAN_XLSX_FMT, _pick_allocation, _server_key, _user_key

MLOB_COLS = ["User ID", "Portfolio Name", "Transaction", "Avg Price",
             "Filled Quantity", "Status", "Server"]

PORTFOLIO_SUMMARY_HEADER = ["Algo", "Server", "Portfolio Executed Users",
                            "No. of Portfolio", "Total Orders", "PnL"]

DEFAULT_PORTFOLIO_PATTERN = "QS"


def read_mlob(source, name=None):
    """Read the Multileg Orders file (Excel or CSV; path or file-like),
    keeping only the columns the analysis needs and the COMPLETE rows."""
    filename = str(name or getattr(source, "name", source)).lower()
    if filename.endswith(".csv"):
        df = pd.read_csv(source, usecols=MLOB_COLS)
    else:
        df = pd.read_excel(source, sheet_name=0, usecols=MLOB_COLS)
    df = df[df["Status"].astype(str).str.strip().str.upper() == "COMPLETE"].copy()
    return df


def portfolio_groups(mlob_df, allocations=None):
    """Collapse the MLOB into one group per (algo, server, user, portfolio):
    COMPLETE order count and PnL (sell value − buy value). This is the
    smallest unit every report — QS or any other pattern — aggregates from."""
    allocations = allocations or {}

    df = mlob_df.rename(columns={
        "User ID": "user", "Server": "server", "Portfolio Name": "portfolio",
    })
    value = (pd.to_numeric(df["Avg Price"], errors="coerce").fillna(0)
             * pd.to_numeric(df["Filled Quantity"], errors="coerce").fillna(0))
    tx = df["Transaction"].astype(str).str.strip().str.upper()
    df = df.assign(
        user=df["user"].astype(str).str.strip(),
        server=df["server"].astype(str).str.strip(),
        portfolio=df["portfolio"].astype(str).str.strip(),
        _sell=value.where(tx == "SELL", 0.0),
        _buy=value.where(tx == "BUY", 0.0),
    )
    # rows without a portfolio name can't be analysed per portfolio
    df = df[~df["portfolio"].str.lower().isin(["", "nan", "none"])]
    grouped = (df.groupby(["user", "server", "portfolio"])
                 .agg(orders=("_sell", "size"), sell=("_sell", "sum"), buy=("_buy", "sum"))
                 .reset_index())

    # server -> algo fallback for users with no MTM entry (same rule as the
    # strikes table: the algo of the server's MTM accounts, most common wins)
    server_algo_votes = {}
    for entries in allocations.values():
        for entry in entries:
            if entry["algo"]:
                server_algo_votes.setdefault(entry["server"], Counter())[entry["algo"]] += 1
    server_algo = {server: votes.most_common(1)[0][0]
                   for server, votes in server_algo_votes.items()}

    algo_cache = {}

    def algo_for(user, server):
        key = (user, server)
        if key not in algo_cache:
            entry = _pick_allocation(allocations, _user_key(user), server)
            algo_cache[key] = (entry["algo"] if entry
                               else server_algo.get(_server_key(server), ""))
        return algo_cache[key]

    return [
        {
            "algo": algo_for(rec.user, rec.server),
            "server": _server_key(rec.server),
            "user_id": rec.user,
            "portfolio": rec.portfolio,
            "orders": int(rec.orders),
            "pnl": float(rec.sell - rec.buy),
        }
        for rec in grouped.itertuples(index=False)
    ]


def _algo_sort_key(algo):
    text = str(algo)
    return (2, "") if text == "" else (0, int(text)) if text.isdigit() else (1, text)


def portfolio_report(groups, pattern=DEFAULT_PORTFOLIO_PATTERN):
    """Aggregate the groups for every portfolio whose name contains `pattern`
    (case-insensitive) into the three-level drill-down
    Algo -> Server -> User. Returns {"pattern", "algos", "total"}:

      algos — per algo: distinct server count, distinct executed users (a
              user on two servers counts once), distinct portfolios, total
              COMPLETE orders, PnL, and its "server_rows" — each server with
              the same metrics plus its per-user "user_rows" (worst PnL
              first);
      total — distinct users / portfolios over the whole selection, summed
              orders and PnL.

    Returns empty "algos" when nothing matches."""
    pat = str(pattern).strip().upper()
    selected = [g for g in groups if pat in g["portfolio"].upper()]

    by_algo = {}
    for g in selected:
        algo = by_algo.setdefault(g["algo"], {
            "servers": {}, "users": set(), "portfolios": set(), "orders": 0, "pnl": 0.0,
        })
        server = algo["servers"].setdefault(g["server"], {
            "users": {}, "portfolios": set(), "orders": 0, "pnl": 0.0,
        })
        user = server["users"].setdefault(g["user_id"], {
            "portfolios": set(), "orders": 0, "pnl": 0.0,
        })
        for target in (algo, server, user):
            target["orders"] += g["orders"]
            target["pnl"] += g["pnl"]
        algo["users"].add(_user_key(g["user_id"]))
        algo["portfolios"].add(g["portfolio"])
        server["portfolios"].add(g["portfolio"])
        user["portfolios"].add(g["portfolio"])

    algos = []
    for algo_name in sorted(by_algo, key=_algo_sort_key):
        algo = by_algo[algo_name]
        server_rows = []
        for server_name in sorted(algo["servers"]):
            server = algo["servers"][server_name]
            user_rows = [
                {"user_id": user_id, "portfolios": len(u["portfolios"]),
                 "orders": u["orders"], "pnl": u["pnl"]}
                for user_id, u in server["users"].items()
            ]
            user_rows.sort(key=lambda u: u["pnl"])
            server_rows.append({
                "server": server_name,
                "users": len(server["users"]),
                "portfolios": len(server["portfolios"]),
                "orders": server["orders"],
                "pnl": server["pnl"],
                "user_rows": user_rows,
            })
        algos.append({
            "algo": algo_name,
            "n_servers": len(algo["servers"]),
            "users": len(algo["users"]),
            "portfolios": len(algo["portfolios"]),
            "orders": algo["orders"],
            "pnl": algo["pnl"],
            "server_rows": server_rows,
        })
    total = {
        "users": len({_user_key(g["user_id"]) for g in selected}),
        "portfolios": len({g["portfolio"] for g in selected}),
        "orders": sum(g["orders"] for g in selected),
        "pnl": sum(g["pnl"] for g in selected),
    }
    return {"pattern": pattern, "algos": algos, "total": total}


def portfolio_json(groups):
    """Compact encoding of the groups for embedding in DOR.html — string
    tables plus index rows [algo, server, user, portfolio, orders, pnl] —
    so the page can aggregate any portfolio pattern client-side."""
    algos, servers, users, portfolios = [], [], [], []
    index = {}

    def idx(table, value):
        key = (id(table), value)
        if key not in index:
            index[key] = len(table)
            table.append(value)
        return index[key]

    rows = [
        [idx(algos, str(g["algo"])), idx(servers, g["server"]),
         idx(users, g["user_id"]), idx(portfolios, g["portfolio"]),
         g["orders"], round(g["pnl"])]
        for g in groups
    ]
    return {"algos": algos, "servers": servers, "users": users,
            "portfolios": portfolios, "rows": rows}


def add_portfolio_sheet(workbook, report):
    """Append the "Portfolio QS" sheet in the three-level layout: one row per
    algo (server COUNT in the Server column), its server rows underneath
    (server name), and each server's user rows under that (user id in the
    users column), closed by a grand total row."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    sheet = workbook.create_sheet("Portfolio QS")
    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    algo_fill = PatternFill("solid", fgColor="B7C7E3")
    server_fill = PatternFill("solid", fgColor="E1E8F5")
    total_fill = PatternFill("solid", fgColor="D1C9E1")

    sheet.append(PORTFOLIO_SUMMARY_HEADER)
    for col_idx, title in enumerate(PORTFOLIO_SUMMARY_HEADER, start=1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = bold
        cell.alignment = center
        sheet.column_dimensions[get_column_letter(col_idx)].width = max(len(title) + 4, 12)

    def write_row(r, values, fill=None, is_bold=False):
        for c, v in enumerate(values, start=1):
            cell = sheet.cell(row=r, column=c, value=v)
            cell.alignment = center
            if fill is not None:
                cell.fill = fill
            if is_bold:
                cell.font = bold
            if c == 6:
                cell.number_format = INDIAN_XLSX_FMT

    r = 2
    for algo_row in report["algos"]:
        write_row(r, [algo_row["algo"], algo_row["n_servers"], algo_row["users"],
                      algo_row["portfolios"], algo_row["orders"], round(algo_row["pnl"])],
                  fill=algo_fill, is_bold=True)
        r += 1
        for server_row in algo_row["server_rows"]:
            write_row(r, ["", server_row["server"], server_row["users"],
                          server_row["portfolios"], server_row["orders"],
                          round(server_row["pnl"])],
                      fill=server_fill, is_bold=True)
            r += 1
            for u in server_row["user_rows"]:
                write_row(r, ["", "", u["user_id"], u["portfolios"], u["orders"],
                              round(u["pnl"])])
                r += 1
    total = report["total"]
    write_row(r, ["Total", "", total["users"], total["portfolios"],
                  total["orders"], round(total["pnl"])], fill=total_fill, is_bold=True)
    sheet.freeze_panes = "A2"
