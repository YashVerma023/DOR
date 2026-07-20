"""
Intraday / Positional segregation report.

Logic
-----
1. Read every account from the Updated Compiled User MTM file.
2. An account is POSITIONAL if its User ID appears in EITHER of the two
   "Combined max loss Calculation" files (23-06 / 20-06); otherwise INTRADAY.
   The same User ID can exist on two servers as two DISTINCT accounts
   (different SERVER + ALGO in the compiled sheet); such ids are matched per
   (User ID, Server) so only the right account is flagged / credited.
3. For a positional account the report's Realized P&L is:
       compiled Realized P&L
     + (Realized PNL + Net Settlement Value) from the 23-06 file (if present)
     + (Realized PNL + Net Settlement Value) from the 20-06 file (if present)
   i.e. summed across BOTH files when the account is in both.
   Intraday accounts keep the plain compiled Realized P&L.
4. Output is a brand-new workbook, grouped per ALGO, and inside each algo split
   into an Int block and a Pos+Int block, each with its own Sub-Total,
   followed by a single overall Grand Total.

This module is import-safe (used by the Streamlit app for the pivot view);
running it directly keeps the original interactive CLI flow.
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ====================== PATHS ======================
# Output is saved in the same folder where this script lives.
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


# ====================== ASK FOR INPUT FILES ======================
def ask_file(label, required):
    while True:
        raw = input(f"  >> {label}\n     Path: ").strip().strip('"').strip("'").strip()
        if not raw:
            if required:
                print("     [!] This file is required - please provide a path.\n")
                continue
            print("     [i] Skipped (no file given).\n")
            return None
        if not os.path.exists(raw):
            print(f"     [!] File not found: {raw}\n")
            if required:
                continue
            again = input("     Type a path to retry, or press Enter to skip: ").strip().strip('"').strip("'")
            if not again:
                print("     [i] Skipped.\n")
                return None
            if os.path.exists(again):
                return again
            print("     [!] Still not found - skipping.\n")
            return None
        return raw


def ask_date(label):
    """Ask for the report date. Accepts it with or without surrounding quotes."""
    while True:
        raw = input(f"  >> {label}\n     Date: ").strip().strip('"').strip("'").strip()
        if raw:
            return raw
        print("     [!] Date is required - please enter it (e.g. 09-07-2026).\n")


# ====================== LOAD DATA ======================
def _server_key(value):
    """Canonical server label for matching accounts across files; a
    missing/NaN server reads as blank."""
    key = str(value or "").strip().upper()
    return "" if key in {"NAN", "NONE", "NA"} else key


def load_combined(source):
    """Read one Combined max loss file (path or file-like) into
    {User ID: [{"server", "addon", "type"}, ...]} — one entry per
    (User ID, Server) row. The same User ID can run on two servers as two
    distinct accounts, so entries are kept per server instead of collapsing
    them onto one id."""
    if source is None:
        return {}
    df = pd.read_excel(source, sheet_name="Sheet1")
    df["User ID"] = df["User ID"].astype(str).str.strip()
    df["Realized PNL"] = pd.to_numeric(df["Realized PNL"], errors="coerce").fillna(0)
    df["Net Settlement Value"] = pd.to_numeric(df["Net Settlement Value"], errors="coerce").fillna(0)
    df["User Type"] = df["User Type"].astype(str).str.strip()
    df["_server"] = df["Server"].map(_server_key) if "Server" in df.columns else ""
    df = df.drop_duplicates(subset=["User ID", "_server"], keep="last")
    out = {}
    for _, r in df.iterrows():
        out.setdefault(r["User ID"], []).append({
            "server": r["_server"],
            "addon": float(r["Realized PNL"] + r["Net Settlement Value"]),
            "type": r["User Type"],
        })
    return out


def _assign_entries(book, user_ids, server_keys):
    """Assign every max-loss entry to exactly ONE compiled row (by position).

    A user id with a single compiled row takes all of its entries (server
    labels between the files can disagree or be blank, so the id alone is
    unambiguous). When the same id appears on several servers — distinct
    accounts that differ in SERVER and ALGO — each entry goes to the row
    whose SERVER matches; an entry with a blank/unmatched server falls back
    to the user's first row, so it is never double-counted."""
    rows_by_uid = {}
    for idx, (uid, server) in enumerate(zip(user_ids, server_keys)):
        rows_by_uid.setdefault(uid, []).append((idx, server))
    assigned = {}
    for uid, entries in book.items():
        rows = rows_by_uid.get(uid)
        if not rows:
            continue
        for entry in entries:
            target = rows[0][0]
            if len(rows) > 1 and entry["server"]:
                target = next((i for i, s in rows if s == entry["server"]), rows[0][0])
            assigned.setdefault(target, []).append(entry)
    return assigned


def prepare_comp(compiled_source, four, one):
    """Read the Updated Compiled User MTM (path or file-like) and enrich it
    with Type / addons / adjusted realized / per-user return columns."""
    comp = pd.read_excel(compiled_source, sheet_name="Sheet1")
    comp["UserID"] = comp["UserID"].astype(str).str.strip()

    for col in ["Realized P&L", "ALLOCATION", "MAX LOSS", "Unrealized P&L"]:
        if col in comp.columns:
            comp[col] = pd.to_numeric(comp[col], errors="coerce").fillna(0)
        else:
            comp[col] = 0
    if "SL HIT/NOT" not in comp.columns:
        comp["SL HIT/NOT"] = 0
    if "Alias" not in comp.columns:
        comp["Alias"] = ""

    # The same UserID can appear on two servers as two DISTINCT accounts
    # (different SERVER and ALGO in the compiled sheet), so max-loss entries
    # are assigned per (UserID, SERVER) row — never to every row of the id.
    user_ids    = comp["UserID"].tolist()
    server_keys = (comp["SERVER"].map(_server_key).tolist()
                   if "SERVER" in comp.columns else [""] * len(comp))
    four_rows = _assign_entries(four, user_ids, server_keys)
    one_rows  = _assign_entries(one, user_ids, server_keys)

    def row_meta(i):
        ones, fours = one_rows.get(i, []), four_rows.get(i, [])
        a1 = sum(e["addon"] for e in ones)
        a4 = sum(e["addon"] for e in fours)
        if not ones and not fours:
            return ("Intraday", "", a4, a1, 0.0)
        utype = (ones or fours)[0]["type"]
        applied = a1 if str(utype).strip().lower() == "noren" else a1 + a4
        return ("Positional", utype, a4, a1, applied)

    meta = [row_meta(i) for i in range(len(comp))]
    comp["Type"]          = [m[0] for m in meta]
    comp["User Type"]     = [m[1] for m in meta]
    comp["Addon 4DTE"]    = [m[2] for m in meta]
    comp["Addon 1DTE"]    = [m[3] for m in meta]
    comp["Addon Applied"] = [m[4] for m in meta]
    comp["AdjRealized"]  = comp["Realized P&L"] + comp["Addon Applied"]
    # Per-user return (Realized / Allocation) — basis for the P5/P95 percentile columns.
    comp["UserReturn"]   = comp.apply(
        lambda r: (r["AdjRealized"] / r["ALLOCATION"]) if r["ALLOCATION"] else 0.0, axis=1
    )
    return comp


def comp_stats(comp):
    n_pos   = int((comp["Type"] == "Positional").sum())
    n_noren = int((comp["User Type"].str.lower() == "noren").sum())
    return {
        "accounts": len(comp),
        "positional": n_pos,
        "intraday": len(comp) - n_pos,
        "noren": n_noren,
    }


def infer_report_date(comp):
    """Best-effort report date (dd-mm-YYYY) from the compiled sheet's Date column."""
    if "Date" in comp.columns:
        for value in comp["Date"]:
            if pd.isna(value):
                continue
            if hasattr(value, "strftime"):
                return value.strftime("%d-%m-%Y")
            text = str(value).strip()
            if text:
                return text.split(" ")[0]
    return ""


# ====================== AGGREGATION ======================
def pct_returns(df_block: pd.DataFrame):
    """5th / 95th percentile of the per-user returns inside a group.

    P5/P95 stand in for min/max while controlling for data aberrations
    (intraday pauses/stops, under/over-funded accounts, suboptimal allocations).
    Users with zero allocation have an undefined return and are excluded.
    Returns (p5, p95).
    """
    rr = df_block.loc[df_block["ALLOCATION"] > 0, "UserReturn"].dropna()
    if rr.empty:
        return (0.0, 0.0)
    return (float(rr.quantile(0.05)), float(rr.quantile(0.95)))


def aggregate(sub: pd.DataFrame):
    rows = []
    for server, g in sub.groupby("SERVER"):
        realized = g["AdjRealized"].sum()
        unreal   = g["Unrealized P&L"].sum()
        alloc    = g["ALLOCATION"].sum()
        p5, p95  = pct_returns(g)
        rows.append({
            "ALGO"      : g["ALGO"].iloc[0],
            "SERVER"    : server,
            "Users"     : len(g),
            "SLHit"     : int((g["SL HIT/NOT"] == 1).sum()),
            "MaxLoss"   : g["MAX LOSS"].sum(),
            "Allocation": alloc,
            "Realized"  : realized,
            "Unrealized": unreal,
            "MTM"       : realized + unreal,
            "Return"    : (realized / alloc) if alloc else 0.0,
            "P95"       : p95,
            "P5"        : p5,
        })
    rows.sort(key=lambda x: str(x["SERVER"]))
    return rows


# ====================== STYLES ======================
HEADERS = ["ALGO", "SERVER", "No. of Users", "No. of SL Hit Users", "MAX LOSS",
           "ALLOCATION", "Realized P&L", "Unrealized P&L", "MTM", "Return %",
           "95%", "5%"]
KEYS    = ["ALGO", "SERVER", "Users", "SLHit", "MaxLoss",
           "Allocation", "Realized", "Unrealized", "MTM", "Return",
           "P95", "P5"]
NCOLS   = 1 + len(HEADERS)   # col A = type label, cols B-M = data

# Column letters (A=1, B=2, ..., M=13)
# B=ALGO  C=SERVER  D=Users  E=SLHit  F=MaxLoss
# G=Allocation  H=Realized  I=Unrealized  J=MTM  K=Return  L=P95  M=P5
KEY_COL   = {"Users": "D", "SLHit": "E", "MaxLoss": "F",
             "Allocation": "G", "Realized": "H", "Unrealized": "I", "MTM": "J"}
MONEY_KEYS = {"MaxLoss", "Allocation", "Realized", "Unrealized", "MTM"}
PCT_KEYS   = {"P95", "P5"}   # 5th/95th percentile of per-user returns (written as values)

# --- Border sides ---
_thin = Side(style="thin",   color="C0C0C0")
_med  = Side(style="medium", color="595959")

def _bdr(l=None, r=None, t=None, b=None):
    return Border(left=l or _thin, right=r or _thin, top=t or _thin, bottom=b or _thin)

inner_bdr    = _bdr()
sub_bdr      = _bdr(t=_med, b=_med)     # medium top + bottom — sub-total reads as separator bar
banner_bdr   = _bdr(t=_med)             # top edge of each algo banner
algo_tot_bdr = _bdr(t=_med, b=_med)    # algo total: medium top + bottom
grand_bdr    = Border(left=_med, right=_med, top=_med, bottom=_med)

# --- Alignments ---
center      = Alignment(horizontal="center", vertical="center")
center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
vertical_a  = Alignment(horizontal="center", vertical="center", text_rotation=90)
left_wrap   = Alignment(horizontal="left", vertical="center", wrap_text=True)

# --- Fills ---
# Only title + header are dark. Every algo cycles through its own light palette pair.
title_fill  = PatternFill("solid", fgColor="1E293B")  # Deep Slate Blue/Gray  (title bar)
header_fill = PatternFill("solid", fgColor="64748B")  # Muted Cool Gray (column headers)
grand_fill  = PatternFill("solid", fgColor="D1C9E1")  # light lavender (grand total)

# Per-algo rotating palette — (data_hex, accent_hex)
#   data_hex   : ~10-12% saturation, used for ALL data rows in the algo
#   accent_hex : slightly more saturated, used for the banner, col-A label, algo-total row
#   sub-total  : derived = accent lightened 60% toward white, so it sits between the
#                data rows (lightest) and the accent, giving it its own distinct shade
ALGO_PALETTE = [
    ("FCE8E6", "E8A898"),  # blush pink
    ("E8F2E4", "98C490"),  # sage green
    ("EAE8F8", "9898D8"),  # periwinkle
    ("E4F0F8", "80BCDC"),  # sky blue
    ("F8F4E4", "C8B868"),  # warm cream / amber
    ("E8F6F2", "78BCA8"),  # seafoam / mint
    ("F8E8EE", "D898B0"),  # rose pink
    ("EEE8F8", "B090D0"),  # soft lilac
    ("E4F6F8", "68B0B8"),  # pale teal
]

def _lighten(hex_color, factor):
    """Lighten a hex colour by blending it `factor` of the way toward white."""
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    r = round(r + (255 - r) * factor)
    g = round(g + (255 - g) * factor)
    b = round(b + (255 - b) * factor)
    return f"{r:02X}{g:02X}{b:02X}"

def _algo_fills(i):
    data_hex, accent_hex = ALGO_PALETTE[i % len(ALGO_PALETTE)]
    sub_hex = _lighten(accent_hex, 0.60)   # sub-total: 60% lighter version of the accent
    return (PatternFill("solid", fgColor=data_hex),
            PatternFill("solid", fgColor=accent_hex),
            PatternFill("solid", fgColor=sub_hex))

white_bold = Font(bold=True, color="FFFFFF")
bold       = Font(bold=True)


# ====================== ROW WRITERS ======================
def _base(cell, fill, bdr, is_bold=False):
    cell.fill      = fill
    cell.border    = bdr
    cell.alignment = center
    if is_bold:
        cell.font = bold


def _algo_val(v):
    """Display algo number as integer when it is a whole float."""
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def write_data_row(ws, row, rowdict, fill, is_bold=False, bdr=None):
    b = bdr or inner_bdr
    for j, key in enumerate(KEYS):
        col  = j + 2
        cell = ws.cell(row, col)
        _base(cell, fill, b, is_bold)
        if key == "Return":
            cell.value         = f"=IF(G{row}=0,0,H{row}/G{row})"
            cell.number_format = "0.00"
        elif key == "ALGO":
            cell.value = _algo_val(rowdict[key])
        elif key in MONEY_KEYS:
            cell.value         = round(float(rowdict[key]))
            cell.number_format = "#,##0"
        elif key in PCT_KEYS:
            cell.value         = round(float(rowdict[key]), 4)
            cell.number_format = "0.00"
        else:
            cell.value = rowdict[key]


def write_subtotal_row(ws, row, data_start, data_end, fill, n_servers, pct=None, is_bold=True, bdr=None):
    """Sub-total row: SUM formulas for all aggregated columns.
    pct: {"P95":.., "P5":..} percentiles over all users in the section."""
    b = bdr or sub_bdr
    for j, key in enumerate(KEYS):
        col  = j + 2
        cell = ws.cell(row, col)
        _base(cell, fill, b, is_bold)
        if key == "ALGO":
            cell.value = "Sub-Total"
        elif key == "SERVER":
            cell.value = n_servers
        elif key in KEY_COL:
            cl             = KEY_COL[key]
            cell.value     = f"=SUM({cl}{data_start}:{cl}{data_end})"
            if key in MONEY_KEYS:
                cell.number_format = "#,##0"
        elif key == "Return":
            cell.value         = f"=IF(G{row}=0,0,H{row}/G{row})"
            cell.number_format = "0.00"
        elif key in PCT_KEYS and pct is not None:
            cell.value         = round(float(pct[key]), 4)
            cell.number_format = "0.00"


def write_total_row(ws, row, label, n_servers, ref_rows, fill, pct=None, is_bold=True, bdr=None):
    """Algo-total or Grand-total: reference (sum) the given ref_rows.
    pct: {"P95":.., "P5":..} percentiles over all users in the algo / whole book."""
    b = bdr or algo_tot_bdr
    for j, key in enumerate(KEYS):
        col  = j + 2
        cell = ws.cell(row, col)
        _base(cell, fill, b, is_bold)
        if key == "ALGO":
            cell.value = label
        elif key == "SERVER":
            cell.value = n_servers
        elif key in KEY_COL:
            cl         = KEY_COL[key]
            formula    = "+".join(f"{cl}{rr}" for rr in ref_rows)
            cell.value = f"={formula}"
            if key in MONEY_KEYS:
                cell.number_format = "#,##0"
        elif key == "Return":
            cell.value         = f"=IF(G{row}=0,0,H{row}/G{row})"
            cell.number_format = "0.00"
        elif key in PCT_KEYS and pct is not None:
            cell.value         = round(float(pct[key]), 4)
            cell.number_format = "0.00"


def apply_outer_border(ws, start_row, end_row, start_col=1, end_col=NCOLS):
    """Overlay a medium border on the outer perimeter of a cell range.
    Preserves all inner borders that are already set."""
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row, col)
            b    = cell.border
            cell.border = Border(
                left   = _med if col == start_col else (b.left   or _thin),
                right  = _med if col == end_col   else (b.right  or _thin),
                top    = _med if row == start_row else (b.top    or _thin),
                bottom = _med if row == end_row   else (b.bottom or _thin),
            )


# ====================== PIVOT ROWS (for on-screen display) ======================
def pivot_rows(comp):
    """The pivot view as a flat list of dicts — same categorisation as the
    workbook's Segregation sheet: per algo, an Int block (intraday accounts)
    and a Pos+Int block (positional accounts), each grouped by server. A
    server running both account types appears in both blocks, split by type.
    Each dict carries a `kind`: data / subtotal / algototal / grandtotal."""

    def block_row(kind, section, algo_label, server, block):
        realized = block["AdjRealized"].sum()
        unreal   = block["Unrealized P&L"].sum()
        alloc    = block["ALLOCATION"].sum()
        p5, p95  = pct_returns(block)
        return {
            "kind": kind, "Section": section,
            "ALGO": algo_label, "SERVER": server,
            "Users": len(block), "SLHit": int((block["SL HIT/NOT"] == 1).sum()),
            "MaxLoss": round(float(block["MAX LOSS"].sum())),
            "Allocation": round(float(alloc)),
            "Realized": round(float(realized)),
            "Unrealized": round(float(unreal)),
            "MTM": round(float(realized + unreal)),
            "Return": round(float(realized / alloc), 2) if alloc else 0.0,
            "P95": round(p95, 2), "P5": round(p5, 2),
        }

    out = []
    algos = sorted(comp["ALGO"].dropna().unique(), key=lambda x: float(x))
    total_n_servers = 0
    for algo in algos:
        algo_df    = comp[comp["ALGO"] == algo]
        algo_label = f"Algo {int(algo) if float(algo).is_integer() else algo}"
        algo_val   = _algo_val(algo)
        for ttype, display_label in [("Intraday", "Int"), ("Positional", "Pos+Int")]:
            sub = algo_df[algo_df["Type"] == ttype]
            if sub.empty:
                continue
            for rowdict in aggregate(sub):
                out.append({
                    "kind": "data", "Section": display_label, "algo_val": algo_val,
                    "ALGO": _algo_val(rowdict["ALGO"]), "SERVER": rowdict["SERVER"],
                    "Users": rowdict["Users"], "SLHit": rowdict["SLHit"],
                    "MaxLoss": round(float(rowdict["MaxLoss"])),
                    "Allocation": round(float(rowdict["Allocation"])),
                    "Realized": round(float(rowdict["Realized"])),
                    "Unrealized": round(float(rowdict["Unrealized"])),
                    "MTM": round(float(rowdict["MTM"])),
                    "Return": round(float(rowdict["Return"]), 2),
                    "P95": round(float(rowdict["P95"]), 2),
                    "P5": round(float(rowdict["P5"]), 2),
                })
            sub_row = block_row("subtotal", display_label, "Sub-Total",
                                sub["SERVER"].nunique(), sub)
            sub_row["algo_val"] = algo_val
            out.append(sub_row)
        algo_n = algo_df["SERVER"].nunique()
        total_row = block_row("algototal", "", f"{algo_label} Total", algo_n, algo_df)
        total_row["algo_val"] = algo_val
        out.append(total_row)
        total_n_servers += algo_n
    grand_row = block_row("grandtotal", "", "Grand Total", total_n_servers, comp)
    grand_row["algo_val"] = ""
    out.append(grand_row)
    return out


# ====================== SLIPPAGE (REALIZED LOSS % vs MAX-LOSS %) ======================
SLIP_THRESHOLD = 0.1   # of allocation — same ratio units as the values themselves
_SLIP_EPS = 1e-9       # float guard so a difference of exactly 0.1 counts


def slippage_rows(comp):
    """One row per SLIPPAGE account, evaluated over ALL accounts with
    ALLOCATION > 0. Both limits are plain RATIOS of the account's allocation
    (positive numbers = loss) — the same unit convention as the pivot's
    Return %, NOT multiplied by 100:

        ML %          = MAX LOSS / ALLOCATION
        Realized ML % = -compiled Realized P&L / ALLOCATION

    An account HAS SLIPPAGE only when its realized loss overshoots the
    configured max-loss by at least 0.1:

        Realized ML % - ML % >= 0.1

    (ML% 1.00 -> Realized 1.09 is NOT slippage; 1.10 is. A profit or a loss
    inside the limit is never slippage.) Sorted worst-first (largest
    Realized ML %)."""
    rows = []
    for _, r in comp.iterrows():
        alloc = float(r["ALLOCATION"])
        if alloc <= 0:
            continue
        ml_pct = float(r["MAX LOSS"]) / alloc
        realized_ml_pct = -float(r["Realized P&L"]) / alloc
        if realized_ml_pct - ml_pct < SLIP_THRESHOLD - _SLIP_EPS:
            continue
        rows.append({
            "ALGO": r["ALGO"],
            "SERVER": r["SERVER"],
            "UserID": r["UserID"],
            "Alias": "" if pd.isna(r["Alias"]) else r["Alias"],
            "Allocation": alloc,
            "MaxLoss": float(r["MAX LOSS"]),
            "Realized": float(r["Realized P&L"]),
            "MLPct": ml_pct,
            "RealizedMLPct": realized_ml_pct,
            "DiffPct": realized_ml_pct - ml_pct,
        })
    rows.sort(key=lambda x: -x["RealizedMLPct"])
    return rows


def slippage_summary(rows, comp):
    """Per algo + overall: total accounts, slippage accounts, and the
    AVERAGE SLIPPAGE of the algo = the mean Realized ML % over the algo's
    slippage accounts (None when the algo has none)."""
    def block(sub, n_accounts):
        return {
            "accounts": n_accounts,
            "slipped": len(sub),
            "avg_slippage": (sum(r["RealizedMLPct"] for r in sub) / len(sub)) if sub else None,
        }

    algos = sorted(comp["ALGO"].dropna().unique(), key=float)
    per_algo = [{"ALGO": _algo_val(algo),
                 **block([r for r in rows if r["ALGO"] == algo],
                         int((comp["ALGO"] == algo).sum()))}
                for algo in algos]
    return per_algo, {"ALGO": "Overall", **block(rows, len(comp))}


def major_slippages(rows, per_algo):
    """Major slippage = a slippage account whose Realized ML % is GREATER
    than its algo's average slippage. Each row carries the algo average for
    display; the worst-first order of `rows` is preserved."""
    avg_by_algo = {s["ALGO"]: s["avg_slippage"] for s in per_algo}
    out = []
    for r in rows:
        avg = avg_by_algo.get(_algo_val(r["ALGO"]))
        if avg is not None and r["RealizedMLPct"] > avg:
            out.append({**r, "AlgoAvgSlippage": avg})
    return out


SLIP_SUMMARY_HEADERS = ["ALGO", "Accounts", "Slippage Accounts", "Avg Slippage %"]
SLIP_DETAIL_HEADERS = ["ALGO", "SERVER", "UserID", "Alias", "ALLOCATION", "MAX LOSS",
                       "Realized P&L", "ML %", "Realized ML %", "Diff %"]
SLIP_MAJOR_HEADERS = SLIP_DETAIL_HEADERS + ["Algo Avg Slippage %"]


def _write_slippage_sheet(wb, comp, report_date):
    """Sheet "Slippage": the per-algo summary on top (accounts / slippage
    accounts / avg slippage %), then every slippage account worst-first with
    the MAJOR rows (Realized ML % above the algo average) highlighted."""
    rows = slippage_rows(comp)
    per_algo, overall = slippage_summary(rows, comp)
    avg_by_algo = {s["ALGO"]: s["avg_slippage"] for s in per_algo}

    ws4 = wb.create_sheet("Slippage")
    n_cols = max(len(SLIP_SUMMARY_HEADERS), len(SLIP_DETAIL_HEADERS))

    ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    tc = ws4.cell(1, 1, f"Slippage — Realized Loss % vs Max Loss % of Allocation  ({report_date})")
    tc.font = Font(bold=True, italic=True, color="FFFFFF", size=13)
    tc.fill = title_fill
    tc.alignment = center
    ws4.row_dimensions[1].height = 22

    def header_row(r, headers):
        for j, h in enumerate(headers, start=1):
            cell = ws4.cell(r, j, h)
            cell.fill = header_fill
            cell.font = white_bold
            cell.alignment = center_wrap
            cell.border = inner_bdr

    r = 2
    header_row(r, SLIP_SUMMARY_HEADERS)
    r += 1
    grand_row_fill = PatternFill("solid", fgColor="D1C9E1")
    for srow in per_algo + [overall]:
        is_overall = srow["ALGO"] == "Overall"
        values = [srow["ALGO"], srow["accounts"], srow["slipped"],
                  round(srow["avg_slippage"], 2) if srow["avg_slippage"] is not None else "—"]
        for j, v in enumerate(values, start=1):
            cell = ws4.cell(r, j, v)
            cell.alignment = center
            cell.border = inner_bdr
            if is_overall:
                cell.fill = grand_row_fill
                cell.font = bold
            if j == 4 and v != "—":
                cell.number_format = "0.00"
        r += 1

    r += 1  # spacer
    detail_header = r
    if rows:
        header_row(r, SLIP_DETAIL_HEADERS)
        r += 1
        major_fill = PatternFill("solid", fgColor="FCE8E6")   # major slippage rows
        for row in rows:
            avg = avg_by_algo.get(_algo_val(row["ALGO"]))
            is_major = avg is not None and row["RealizedMLPct"] > avg
            values = [_algo_val(row["ALGO"]), row["SERVER"], row["UserID"], row["Alias"],
                      round(row["Allocation"]), round(row["MaxLoss"]), round(row["Realized"]),
                      round(row["MLPct"], 2), round(row["RealizedMLPct"], 2),
                      round(row["DiffPct"], 2)]
            for j, v in enumerate(values, start=1):
                cell = ws4.cell(r, j, v)
                cell.alignment = center
                cell.border = inner_bdr
                if is_major:
                    cell.fill = major_fill
                if j in (5, 6, 7):
                    cell.number_format = "#,##0"
                elif j in (8, 9, 10):
                    cell.number_format = "0.00"
            r += 1
    else:
        ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
        nc = ws4.cell(r, 1, "-- No slippage today --")
        nc.alignment = center
        nc.font = Font(italic=True, color="595959")
        r += 1

    note_row = r + 1
    ws4.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=n_cols)
    nc = ws4.cell(
        note_row, 1,
        "*ML % = MAX LOSS / ALLOCATION; Realized ML % = |compiled Realized P&L| / ALLOCATION "
        "— plain ratios of allocation, same convention as Return %. An account has slippage "
        "only when Realized ML % exceeds ML % by at least 0.1 (1.00 → 1.09 is not slippage; "
        "1.10 is). Avg Slippage = average Realized ML % of the algo's slippage accounts; "
        "highlighted rows are MAJOR slippages — accounts above their algo's average."
    )
    nc.font = Font(italic=True, color="595959", size=9)
    nc.alignment = left_wrap
    ws4.row_dimensions[note_row].height = 40

    ws4.freeze_panes = f"A{detail_header + 1}"
    slip_widths = [7, 10, 20, 22, 13, 13, 14, 9, 13, 9]
    for j, w in enumerate(slip_widths, start=1):
        ws4.column_dimensions[get_column_letter(j)].width = w


# ====================== WORKBOOK ======================
def add_segregation_sheets(wb, comp, report_date):
    """Append the Segregation pivot + Raw_Data_Per_User + Worst 10%ile sheets
    to an existing openpyxl workbook (used to combine reports in one file)."""
    ws = wb.create_sheet("Segregation")

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
    tc = ws.cell(1, 1, f"Algo & Server wise Realized P&L  —  Int / Pos+Int  ({report_date})")
    tc.font      = Font(bold=True, italic=True, color="FFFFFF", size=13)
    tc.fill      = title_fill
    tc.alignment = center

    # Header row
    HEADER_ROW = 2
    ws.cell(HEADER_ROW, 1, "Type")
    for j, h in enumerate(HEADERS):
        ws.cell(HEADER_ROW, j + 2, h)
    for c in range(1, NCOLS + 1):
        cell           = ws.cell(HEADER_ROW, c)
        cell.fill      = header_fill
        cell.font      = white_bold
        cell.alignment = center_wrap
        cell.border    = inner_bdr

    # ---- Build sheet ----
    r               = HEADER_ROW + 1
    algo_total_rows = []    # row numbers of each algo-total row (for grand total formula)
    total_n_servers = 0

    algos = sorted(comp["ALGO"].dropna().unique(), key=lambda x: float(x))

    for i, algo in enumerate(algos):
        algo_df        = comp[comp["ALGO"] == algo]
        algo_label     = f"Algo {int(algo) if float(algo).is_integer() else algo}"
        d_fill, a_fill, s_fill = _algo_fills(i)   # data + accent + sub-total fills for this algo

        # ---- Algo banner (accent colour, medium top border) ----
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=NCOLS)
        bc           = ws.cell(r, 2, algo_label)
        bc.fill      = a_fill
        bc.font      = Font(bold=True, color="1F2D3D", size=11)
        bc.alignment = center
        for c in range(1, NCOLS + 1):
            cell        = ws.cell(r, c)
            cell.fill   = a_fill
            cell.border = banner_bdr
        ws.row_dimensions[r].height = 20
        r += 1

        section_sub_rows = []

        for ttype, display_label in [("Intraday", "Int"), ("Positional", "Pos+Int")]:
            sub = algo_df[algo_df["Type"] == ttype]
            if sub.empty:
                continue

            rows        = aggregate(sub)
            block_start = r

            for rowdict in rows:
                write_data_row(ws, r, rowdict, d_fill)
                ws.row_dimensions[r].height = 18
                r += 1

            data_end = r - 1
            sub_row  = r
            # sub-total uses its own 60%-lighter accent shade so it reads distinct from data rows
            sub_p5, sub_p95 = pct_returns(sub)
            write_subtotal_row(ws, r, block_start, data_end, s_fill, n_servers=len(rows),
                               pct={"P95": sub_p95, "P5": sub_p5})
            ws.row_dimensions[r].height = 18
            r += 1

            # Section label col A — merged, rotated, accent colour
            ws.merge_cells(start_row=block_start, start_column=1, end_row=sub_row, end_column=1)
            lc           = ws.cell(block_start, 1, display_label)
            lc.fill      = a_fill
            lc.font      = Font(bold=True, color="1F2D3D")
            lc.alignment = vertical_a
            for rr in range(block_start, sub_row + 1):
                ws.cell(rr, 1).fill   = a_fill
                ws.cell(rr, 1).border = inner_bdr

            # Outer medium border boxes the whole Int / Pos+Int section
            apply_outer_border(ws, block_start, sub_row)

            section_sub_rows.append(sub_row)

        # Algo-level server count is DISTINCT across Int + Pos+Int: a server that runs
        # both types appears in each section but should be counted once for the algo.
        algo_n_servers = algo_df["SERVER"].nunique()

        # ---- Algo Total — col A:B merged & centered (no orphan empty cell) ----
        algo_p5, algo_p95 = pct_returns(algo_df)
        write_total_row(ws, r, f"{algo_label} Total", algo_n_servers, section_sub_rows, a_fill,
                        pct={"P95": algo_p95, "P5": algo_p5})
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        mc           = ws.cell(r, 1)          # top-left cell owns the merged range
        mc.value     = f"{algo_label} Total"
        mc.fill      = a_fill
        mc.font      = bold
        mc.alignment = center
        mc.border    = algo_tot_bdr
        ws.row_dimensions[r].height = 20
        apply_outer_border(ws, r, r)
        algo_total_rows.append(r)
        total_n_servers += algo_n_servers
        r += 1

        r += 1   # blank spacer row between algos

    # ---- Grand total ----
    grand_p5, grand_p95 = pct_returns(comp)
    write_total_row(ws, r, "Grand Total", total_n_servers, algo_total_rows, grand_fill,
                    pct={"P95": grand_p95, "P5": grand_p5}, bdr=grand_bdr)
    gc           = ws.cell(r, 1, "Total")
    gc.fill      = grand_fill
    gc.font      = bold
    gc.alignment = center
    gc.border    = grand_bdr

    # ---- P&L sign colouring: Realized (H) / Unrealized (I) / MTM (J) ----
    # profit green, loss red — conditional formatting so the live formulas
    # keep recolouring when the sheet is edited
    pnl_range = f"H{HEADER_ROW + 1}:J{r}"
    ws.conditional_formatting.add(pnl_range, CellIsRule(
        operator="greaterThan", formula=["0"], font=Font(color="15803D")))
    ws.conditional_formatting.add(pnl_range, CellIsRule(
        operator="lessThan", formula=["0"], font=Font(color="DC2626")))

    # ---- Footnote explaining the P5/P95 (95% / 5%) columns ----
    note_row = r + 2
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=NCOLS)
    nc = ws.cell(
        note_row, 1,
        "*P5/P95 are the 5th and 95th percentile, and are used for representing min/max "
        "respectively. This is done to control for data aberrations due to exceptions like "
        "intraday account pause/stops (tech errors/funding issue/broker issue etc), returns "
        "due to underfunded or overfunded accounts, or statistically abnormal returns due to "
        "suboptimal allocations etc."
    )
    nc.font      = Font(italic=True, color="595959", size=9)
    nc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[note_row].height = 46

    # ---- Finish formatting ----
    ws.freeze_panes = f"A{HEADER_ROW + 1}"
    ws.column_dimensions["A"].width = 5
    widths = [7, 9, 12, 18, 13, 13, 14, 15, 13, 10, 9, 9]
    for j, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(j + 2)].width = w
    ws.row_dimensions[1].height = 22

    _write_raw_sheet(wb, comp)
    _write_worst_sheet(wb, comp, algos, report_date)
    _write_slippage_sheet(wb, comp, report_date)


def build_workbook(comp, report_date):
    """Build the full styled workbook (Segregation pivot + Raw_Data_Per_User +
    Worst 10%ile) and return it."""
    wb = Workbook()
    wb.remove(wb.active)
    add_segregation_sheets(wb, comp, report_date)
    return wb


# ====================== RAW PER-USER DATA SHEET ======================
def _write_raw_sheet(wb, comp):
    comp = comp.copy()
    comp["MTM_calc"] = comp["AdjRealized"] + comp["Unrealized P&L"]
    comp["SL Hit"]   = (comp["SL HIT/NOT"] == 1).astype(int)

    raw_cols = [
        ("Type",                 "Type"),
        ("User Type",            "User Type"),
        ("ALGO",                 "ALGO"),
        ("SERVER",               "SERVER"),
        ("UserID",               "UserID"),
        ("Alias",                "Alias"),
        ("SL Hit",               "SL Hit"),
        ("MAX LOSS",             "MAX LOSS"),
        ("ALLOCATION",           "ALLOCATION"),
        ("Compiled Realized P&L","Realized P&L"),
        ("Addon 4DTE",           "Addon 4DTE"),
        ("Addon 1DTE",           "Addon 1DTE"),
        ("Addon Applied",        "Addon Applied"),
        ("Realized P&L (Final)", "AdjRealized"),
        ("Unrealized P&L",       "Unrealized P&L"),
        ("MTM",                  "MTM_calc"),
    ]
    raw = comp[[src for _, src in raw_cols]].copy()
    raw.columns = [hdr for hdr, _ in raw_cols]
    raw = raw.sort_values(["ALGO", "Type", "SERVER", "UserID"]).reset_index(drop=True)

    ws2 = wb.create_sheet("Raw_Data_Per_User")
    for j, hdr in enumerate(raw.columns, start=1):
        cell           = ws2.cell(1, j, hdr)
        cell.fill      = header_fill
        cell.font      = white_bold
        cell.alignment = center_wrap
        cell.border    = inner_bdr

    money_cols = {"MAX LOSS", "ALLOCATION", "Compiled Realized P&L",
                  "Addon 4DTE", "Addon 1DTE", "Addon Applied",
                  "Realized P&L (Final)", "Unrealized P&L", "MTM"}
    for i, rec in enumerate(raw.itertuples(index=False), start=2):
        for j, (hdr, val) in enumerate(zip(raw.columns, rec), start=1):
            cell        = ws2.cell(i, j, val)
            cell.border = inner_bdr
            if hdr in money_cols and isinstance(val, (int, float)):
                cell.value         = float(val)
                cell.number_format = "#,##0"
            if hdr in ("Type", "ALGO", "SERVER", "SL Hit"):
                cell.alignment = center

    ws2.freeze_panes = "A2"
    raw_widths = [11, 7, 9, 12, 22, 7, 12, 12, 18, 18, 18, 15, 13]
    for j, w in enumerate(raw_widths, start=1):
        ws2.column_dimensions[get_column_letter(j)].width = w


# ====================== WORST 10%ILE EXCEPTION SHEET ======================
def worst_reason(sl_hit, alloc, ret):
    """Best-guess reason for a poor performer — editable in Excel afterwards."""
    if sl_hit == 1:
        return "SL Hit"
    if alloc <= 0:
        return "Zero/low allocation"
    if ret < 0:
        return "Negative return"
    return "Low relative return"


def worst_decile(df_block):
    """Users at/below the 10th percentile of per-user return within the block.
    Only accounts with allocation > 0 have a defined return. Sorted worst-first."""
    valid = df_block[df_block["ALLOCATION"] > 0].copy()
    if valid.empty:
        return valid
    thr = valid["UserReturn"].quantile(0.10)
    return valid[valid["UserReturn"] <= thr].sort_values("UserReturn")


WORST_HEADERS = ["ALGO", "Type", "SERVER", "UserID", "Alias", "SL Hit",
                 "ALLOCATION", "Realized P&L (Final)", "Return %",
                 "Algo Avg Return %", "Reason"]
W_NCOLS = len(WORST_HEADERS)
W_MONEY = {"ALLOCATION", "Realized P&L (Final)"}
W_PCT   = {"Return %", "Algo Avg Return %"}


def group_return(df_block):
    """Section-level return = sum(Realized) / sum(Allocation) over the whole
    Algo + Int/Pos group — matches that section's Return % sub-total on sheet 1."""
    alloc = df_block["ALLOCATION"].sum()
    return (df_block["AdjRealized"].sum() / alloc) if alloc else 0.0


def _write_worst_sheet(wb, comp, algos, report_date):
    comp = comp.copy()
    comp["SL Hit"] = (comp["SL HIT/NOT"] == 1).astype(int)

    ws3 = wb.create_sheet("Worst 10%ile")

    # Title row
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=W_NCOLS)
    wt = ws3.cell(1, 1, f"Algo-wise Exception Report  —  Worst 10%ile Performers  ({report_date})")
    wt.font      = Font(bold=True, italic=True, color="FFFFFF", size=13)
    wt.fill      = title_fill
    wt.alignment = center

    # Header row
    W_HEADER_ROW = 2
    for j, h in enumerate(WORST_HEADERS, start=1):
        cell           = ws3.cell(W_HEADER_ROW, j, h)
        cell.fill      = header_fill
        cell.font      = white_bold
        cell.alignment = center_wrap
        cell.border    = inner_bdr

    wr = W_HEADER_ROW + 1
    for i, algo in enumerate(algos):
        algo_df    = comp[comp["ALGO"] == algo]
        algo_label = f"Algo {int(algo) if float(algo).is_integer() else algo}"
        d_fill, a_fill, s_fill = _algo_fills(i)

        for ttype, display_label in [("Intraday", "Int"), ("Positional", "Pos")]:
            section = algo_df[algo_df["Type"] == ttype]
            worst   = worst_decile(section)
            if worst.empty:
                continue
            algo_avg_ret = group_return(section)   # same value as sheet-1 sub-total Return %

            # Section banner: "Algo X  —  Int / Pos"
            ws3.merge_cells(start_row=wr, start_column=1, end_row=wr, end_column=W_NCOLS)
            bc           = ws3.cell(wr, 1, f"{algo_label}  —  {display_label}")
            bc.fill      = a_fill
            bc.font      = Font(bold=True, color="1F2D3D", size=11)
            bc.alignment = Alignment(horizontal="left", vertical="center")
            for c in range(1, W_NCOLS + 1):
                ws3.cell(wr, c).fill   = a_fill
                ws3.cell(wr, c).border = banner_bdr
            ws3.row_dimensions[wr].height = 20
            wr += 1

            disp = worst[["ALGO", "SERVER", "UserID", "Alias", "SL Hit",
                          "ALLOCATION", "AdjRealized", "UserReturn"]]
            for _, row in disp.iterrows():
                sl    = int(row["SL Hit"])
                alloc = float(row["ALLOCATION"])
                ret   = float(row["UserReturn"])
                values = [
                    _algo_val(row["ALGO"]),
                    display_label,
                    row["SERVER"],
                    row["UserID"],
                    row["Alias"],
                    sl,
                    round(alloc),
                    round(float(row["AdjRealized"])),
                    round(ret, 4),
                    round(algo_avg_ret, 4),
                    worst_reason(sl, alloc, ret),
                ]
                for j, val in enumerate(values, start=1):
                    hdr  = WORST_HEADERS[j - 1]
                    cell = ws3.cell(wr, j, val)
                    cell.fill   = d_fill
                    cell.border = inner_bdr
                    if hdr in W_MONEY:
                        cell.number_format = "#,##0"
                        cell.alignment     = center
                    elif hdr in W_PCT:
                        cell.number_format = "0.00"
                        cell.alignment     = center
                    elif hdr in ("Alias", "Reason"):
                        cell.alignment = left_wrap
                    else:
                        cell.alignment = center
                ws3.row_dimensions[wr].height = 18
                wr += 1

    if wr == W_HEADER_ROW + 1:   # nothing qualified
        ws3.merge_cells(start_row=wr, start_column=1, end_row=wr, end_column=W_NCOLS)
        nc = ws3.cell(wr, 1, "No exception users found.")
        nc.alignment = center
        nc.font      = Font(italic=True, color="595959")
        wr += 1

    # Footnote
    w_note_row = wr + 1
    ws3.merge_cells(start_row=w_note_row, start_column=1, end_row=w_note_row, end_column=W_NCOLS)
    wn = ws3.cell(
        w_note_row, 1,
        "*Worst 10%ile = users at or below the 10th percentile of per-user return "
        "(Realized P&L (Final) / ALLOCATION) within each Algo + Int/Pos group; only "
        "accounts with a positive allocation are ranked. The Reason column is "
        "auto-inferred (SL Hit / allocation / negative-return review) and can be edited."
    )
    wn.font      = Font(italic=True, color="595959", size=9)
    wn.alignment = left_wrap
    ws3.row_dimensions[w_note_row].height = 46

    ws3.freeze_panes = f"A{W_HEADER_ROW + 1}"
    worst_widths = [7, 8, 12, 20, 22, 8, 14, 18, 10, 15, 44]
    for j, w in enumerate(worst_widths, start=1):
        ws3.column_dimensions[get_column_letter(j)].width = w
    ws3.row_dimensions[1].height = 22


# ====================== CLI ======================
def main():
    print("=" * 60)
    print("  SEGREGATION REPORT - INPUTS")
    print("=" * 60)
    report_date   = ask_date("Enter the report date (e.g. 09-07-2026):")
    compiled_path = ask_file("Path to the Updated Compiled User MTM file:", required=True)

    print("\n" + "=" * 60)
    print("  COMBINED MAX LOSS - FILE UPLOAD")
    print("  (Both files optional, but at least the 1DTE file is needed.)")
    print("=" * 60)
    path_4dte = ask_file("Upload 4DTE Combined max loss file (press Enter to skip):", required=False)
    path_1dte = ask_file("Upload 1DTE Combined max loss file:", required=True)

    print("\nReading source files...")
    four = load_combined(path_4dte)
    one  = load_combined(path_1dte)
    comp = prepare_comp(compiled_path, four, one)

    stats = comp_stats(comp)
    print(f"  Files loaded -> 4DTE: {'yes' if four else 'NO'} | 1DTE: {'yes' if one else 'NO'}")
    print(f"  Accounts: {stats['accounts']}  |  Positional: {stats['positional']}  |  Intraday: {stats['intraday']}")
    print(f"  Positional breakdown -> Noren: {stats['noren']} | Non-Noren: {stats['positional'] - stats['noren']}")

    print("Writing raw per-user data sheet...")
    print("Writing worst 10%ile exception sheet...")
    wb = build_workbook(comp, report_date)

    output_path = os.path.join(SCRIPT_DIR, f"Segregated_Int_Pos_MTM_{report_date}.xlsx")
    wb.save(output_path)

    py_realized = comp["AdjRealized"].sum()
    py_mtm      = (comp["AdjRealized"] + comp["Unrealized P&L"]).sum()
    print(f"\nDone. Report saved at:\n  {output_path}")
    print(f"  Grand Total -> Users: {stats['accounts']}  Realized P&L: {round(py_realized):,}  MTM: {round(py_mtm):,}")


if __name__ == "__main__":
    main()
