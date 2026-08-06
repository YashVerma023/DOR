"""
Daily summary engine — the Int / Pos+Int pivot behind DOR.html and the Excel
workbook.

Everything is derived from ONE flat table, the "MTM Data" sheet: the compiled
User MTM's own columns, plus Running Type / Running Days from the All User
sheet, plus MTM computed per account. Reading down that sheet reproduces the
pivot exactly, which is the point of it existing — there is no separate
segregation pass to reconcile against.

Logic
-----
1. Read every account from the Updated Compiled User MTM file.
2. Classification comes from the "All User Details" sheet (tab "Main"), NOT
   from any max-loss file. That sheet states each account's `Running Type`
   (INT / POS) and `Running Days` (0DTE / 1DTE-0DTE / DAILY):

     * on upload, a row is DROPPED when any of `server`, `Running Type` or
       `Running Days` reads "DLR ACC" or "NOT RUNNING" — dealer and stopped
       accounts are not part of any report, whatever the DTE;
     * the selected DTE then narrows the surviving rows to the accounts that
       actually run that day:
           0DTE — every running account
           1DTE — Running Days == "1DTE/0DTE"
           4DTE — Running Days == "DAILY"
     * an in-scope account takes its type from `Running Type`:
           INT -> Intraday   (block label "Int")
           POS -> Positional (block label "Pos+Int")

   Accounts are matched to the compiled MTM on the canonical User ID alone
   (leading zeros stripped, case-insensitive) — `userId` is unique in the
   sheet, so the server is not needed to disambiguate.
3. MTM = Realized P&L + Unrealized P&L, computed per account. The compiled
   sheet ships an MTM column but it is only populated where Unrealized is
   non-zero (577 of 633 rows in the 30-07-2026 file), so it is overwritten,
   never read. MTM % = MTM / ALLOCATION, so the percentage always sits on the
   money column beside it.
   MAX LOSS comes from the compiled MTM's own column; the All User `max_loss`
   is carried through as a reference column only and drives nothing.
   The Combined Max Loss files are no longer an input — there is no addon.
4. A compiled account that the DTE scope does not cover — filtered out, or
   absent from the sheet entirely — is typed "Unclassified" and reported in
   its own section at the BOTTOM of the pivot, outside the per-algo blocks.
   Each Algo Total therefore covers only that algo's classified accounts, and
   the Grand Total = the algo totals + the Unclassified sub-total. Normally
   the All User sheet covers everyone and this section is empty.
5. Output is a brand-new workbook, grouped per ALGO, and inside each algo split
   into an Int block and a Pos+Int block, each with its own Sub-Total,
   followed by the Unclassified section (when non-empty) and a single overall
   Grand Total.

This module is import-safe (used by the Streamlit app for the pivot view);
running it directly keeps the original interactive CLI flow.
"""

import json
import os

import pandas as pd

from tradevalue import INDIAN_XLSX_FMT, _user_key
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ====================== PATHS ======================
# Output is saved in the same folder where this script lives.
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


# ====================== ASK FOR INPUT FILES ======================
def ask_file(label, required):
    while True:
        raw = input(f"  >> {label}\n     Path: ").strip().strip('"').strip("'").strip()
        if not raw:
            if required:
                print("     [!] This file is required - please provide a path.\n")
                continue
            print("     [i] Skipped (no file given).\n")
            return None
        if not os.path.exists(raw):
            print(f"     [!] File not found: {raw}\n")
            if required:
                continue
            again = input("     Type a path to retry, or press Enter to skip: ").strip().strip('"').strip("'")
            if not again:
                print("     [i] Skipped.\n")
                return None
            if os.path.exists(again):
                return again
            print("     [!] Still not found - skipping.\n")
            return None
        return raw


def ask_date(label):
    """Ask for the report date. Accepts it with or without surrounding quotes."""
    while True:
        raw = input(f"  >> {label}\n     Date: ").strip().strip('"').strip("'").strip()
        if raw:
            return raw
        print("     [!] Date is required - please enter it (e.g. 09-07-2026).\n")


# ====================== LOAD DATA ======================
def _server_key(value):
    """Canonical server label for matching accounts across files; a
    missing/NaN server reads as blank."""
    key = str(value or "").strip().upper()
    return "" if key in {"NAN", "NONE", "NA"} else key


# ---------------------------------------------------------------------------
# All User Details sheet — the single source of Int / Pos classification
# ---------------------------------------------------------------------------

# The workbook carries a dozen tabs; the account master is "Main".
ALL_USER_SHEET = "Main"

# Only these six columns are read. `max_loss` is a reference column: MAX LOSS
# for every calculation comes from the compiled MTM's own column.
ALL_USER_COLS = ["userId", "server", "algo", "max_loss",
                 "Running Type", "Running Days"]

# A row whose server / Running Type / Running Days reads either of these is a
# dealer or stopped account: dropped on upload, for every DTE.
ALL_USER_DROP_VALUES = {"DLR ACC", "NOT RUNNING"}

# Which Running Days each DTE covers. The sheet's Running Days states the days
# an account RUNS ON, so the scopes are CUMULATIVE, not one label each:
#
#     DAILY      runs every day        -> in scope for 0DTE, 1DTE and 4DTE
#     1DTE/0DTE  runs on 1DTE and 0DTE -> in scope for 0DTE and 1DTE
#     0DTE       runs on expiry only   -> in scope for 0DTE
#
# Matching "1DTE" against the literal string "1DTE/0DTE" alone was wrong: it
# dropped every DAILY account from a 1DTE report even though a DAILY account
# trades that day. The 05-08-2026 MTM settles it — all 374 of its matched
# accounts are 1DTE/0DTE (181) or DAILY (193), and not one is 0DTE-only.
# None = every running account.
DTE_RUNNING_DAYS = {
    "0DTE": None,
    "1DTE": {"1DTE/0DTE", "DAILY"},
    "4DTE": {"DAILY"},
}
DTE_OPTIONS = list(DTE_RUNNING_DAYS)

# Running Type -> the Type value carried on the compiled rows. "Positional"
# still displays as "Pos+Int": a POS account runs intraday as well.
RUNNING_TYPE_TO_TYPE = {"INT": "Intraday", "POS": "Positional"}

# An account the DTE scope does not cover, or that the sheet never listed.
UNCLASSIFIED = "Unclassified"


def _text(value):
    """Canonical comparison form for the sheet's label columns."""
    return str(value if value is not None else "").strip().upper()


def read_all_users(source, name=None):
    """Read the All User Details workbook's "Main" tab and drop the dealer /
    stopped accounts.

    Returns the surviving rows as a DataFrame with the six required columns,
    plus `match_key` (the canonical user id used to match the compiled MTM)
    and the upper-cased `server_key` / `running_type` / `running_days` used by
    the DTE filter. Those helper names deliberately carry NO leading
    underscore — DataFrame.itertuples() renames such columns to positional
    `_1` / `_2` placeholders, which silently breaks attribute access.

    The drop test is applied to `server`, `Running Type` AND `Running Days`:
    any one of them reading "DLR ACC" or "NOT RUNNING" removes the row. It is
    deliberately not an all-three test — a row that is stopped in one column
    but stale in another is still a row no report should classify."""
    if source is None:
        return None
    filename = str(name or getattr(source, "name", "") or "").lower()
    if filename.endswith(".csv"):
        df = pd.read_csv(source)
    else:
        df = pd.read_excel(source, sheet_name=ALL_USER_SHEET)

    missing = [c for c in ALL_USER_COLS if c not in df.columns]
    if missing:
        raise ValueError(
            f"All User sheet \"{ALL_USER_SHEET}\" is missing the column(s): "
            + ", ".join(missing)
        )

    out = df[ALL_USER_COLS].copy()
    out["userId"] = out["userId"].astype(str).str.strip()
    out["max_loss"] = pd.to_numeric(out["max_loss"], errors="coerce")
    out["match_key"] = out["userId"].map(_user_key)
    out["server_key"] = out["server"].map(_server_key)
    out["running_type"] = out["Running Type"].map(_text)
    out["running_days"] = out["Running Days"].map(_text)

    drop = (
        out["server_key"].isin(ALL_USER_DROP_VALUES)
        | out["running_type"].isin(ALL_USER_DROP_VALUES)
        | out["running_days"].isin(ALL_USER_DROP_VALUES)
    )
    out = out[~drop]
    # a blank id can never be matched, and a repeated id would make the type
    # ambiguous — keep the first occurrence rather than guessing
    out = out[out["match_key"] != ""]
    return out.drop_duplicates(subset=["match_key"], keep="first").reset_index(drop=True)


def all_user_reference(all_users):
    """{canonical user id: {"running_type", "running_days", "algo", "server"}}
    over the rows that SURVIVED the drop.

    A dropped row is gone, so it is not in here and never comes back — the
    "unclassified" sheet reports any id it cannot find as "Not Found", which
    covers a dropped row and an id absent from the sheet alike. The only
    accounts that resolve are the ones that are genuinely running but on a
    different DTE."""
    if all_users is None:
        return {}
    return {
        rec.match_key: {
            "running_type": rec.running_type,
            "running_days": rec.running_days,
            "algo": rec.algo,
            "server": rec.server_key,
        }
        for rec in all_users.itertuples(index=False)
    }


def dte_scope(all_users, dte):
    """{canonical user id: {"type", "running_type", "running_days", "server",
    "algo", "max_loss"}} for the accounts the given DTE covers.

    0DTE takes every running account; 1DTE takes Running Days "1DTE/0DTE";
    4DTE takes "DAILY". An account whose Running Type is neither INT nor POS
    is left out of the map, so it reports as Unclassified rather than being
    forced into a block the sheet does not support."""
    if all_users is None:
        return {}
    days = DTE_RUNNING_DAYS.get(str(dte).strip().upper())
    rows = all_users if days is None else all_users[all_users["running_days"].isin(days)]
    scope = {}
    for rec in rows.itertuples(index=False):
        user_type = RUNNING_TYPE_TO_TYPE.get(rec.running_type)
        if user_type is None:
            continue
        scope[rec.match_key] = {
            "type": user_type,
            "running_type": rec.running_type,
            "running_days": rec.running_days,
            "server": rec.server_key,
            "algo": rec.algo,
            "max_loss": (None if pd.isna(rec.max_loss) else float(rec.max_loss)),
            # the id as the All User sheet writes it — shown next to the MTM's
            # own id so an alias match is visible rather than invisible
            "user_id": rec.userId,
        }
    return scope


# ---------------------------------------------------------------------------
# User aliases — accounts whose MTM id differs from their All User id
# ---------------------------------------------------------------------------

ALIAS_FILE = os.path.join(SCRIPT_DIR, "user_aliases.json")

# column names accepted when the map is uploaded as a CSV / Excel table
_ALIAS_SOURCE_COLS = ("all user id", "alluserid", "all_user_id", "userid",
                      "user id", "all user")
_ALIAS_TARGET_COLS = ("mtm id", "mtmid", "mtm_id", "mtm userid", "mtm user id",
                      "compiled id", "alias")


def _alias_pairs_from_frame(df):
    """(All User id, MTM id) pairs from an uploaded two-column table."""
    by_norm = {str(c).strip().lower(): c for c in df.columns}
    src = next((by_norm[c] for c in _ALIAS_SOURCE_COLS if c in by_norm), None)
    tgt = next((by_norm[c] for c in _ALIAS_TARGET_COLS if c in by_norm), None)
    if src is None or tgt is None:
        # unnamed / oddly named headers: fall back to the first two columns
        if len(df.columns) < 2:
            raise ValueError(
                "alias table needs two columns: All User id, MTM id")
        src, tgt = df.columns[0], df.columns[1]
    return list(zip(df[src], df[tgt]))


def load_user_aliases(source=None, name=None):
    """{MTM user key: All User user key} for the accounts whose two files
    disagree on the id.

    The stored map is written the way the desk keeps it — All User id first,
    MTM id second — and is REVERSED here, because the lookup direction is
    "this compiled row calls itself XLDH142, who is that in the All User
    sheet?".

    `source` is an optional uploaded file (JSON, CSV or Excel) that REPLACES
    the on-disk user_aliases.json for that run; with no source and no file on
    disk the map is empty and nothing changes. Entries that resolve to
    themselves are dropped (harmless no-ops, e.g. the zero-padded numeric
    ids the canonical key already unifies), and an MTM id claimed by two
    different All User ids is dropped as ambiguous rather than guessed."""
    raw = {}
    if source is not None:
        filename = str(name or getattr(source, "name", "") or "").lower()
        if filename.endswith((".csv", ".xlsx", ".xlsm", ".xls")):
            df = (pd.read_csv(source) if filename.endswith(".csv")
                  else pd.read_excel(source, sheet_name=0))
            pairs = _alias_pairs_from_frame(df)
        else:
            data = source.read() if hasattr(source, "read") else source
            if isinstance(data, bytes):
                data = data.decode("utf-8-sig", errors="replace")
            pairs = list(json.loads(data).items())
        raw = dict(pairs)
    elif os.path.exists(ALIAS_FILE):
        with open(ALIAS_FILE, encoding="utf-8") as fh:
            raw = json.load(fh)

    out, claimed = {}, {}
    for all_user_id, mtm_id in raw.items():
        # the JSON carries a "_comment" block for the humans editing it
        if str(all_user_id).startswith("_") or isinstance(mtm_id, (list, dict)):
            continue
        src_key, mtm_key = _user_key(all_user_id), _user_key(mtm_id)
        if not src_key or not mtm_key or src_key == mtm_key:
            continue
        if mtm_key in claimed and claimed[mtm_key] != src_key:
            out.pop(mtm_key, None)      # two owners — never guess
            continue
        claimed[mtm_key] = src_key
        out[mtm_key] = src_key
    return out


def scope_stats(all_users, dte):
    """(accounts in scope, positional, intraday) for the DTE — the counts the
    dashboard shows next to the uploader, before any MTM is matched."""
    scope = dte_scope(all_users, dte)
    positional = sum(1 for v in scope.values() if v["type"] == "Positional")
    return len(scope), positional, len(scope) - positional


def read_compiled(source, name=None):
    """Read a compiled User MTM into a DataFrame — CSV or Excel, decided by
    the file name (the uploaders accept both)."""
    if name and str(name).lower().endswith(".csv"):
        return pd.read_csv(source)
    return pd.read_excel(source, sheet_name="Sheet1")


def prepare_comp(compiled_source, scope=None, name=None, aliases=None):
    """Read the Updated Compiled User MTM (path, file-like or an already-read
    DataFrame) and enrich it with the Type / realized / per-user return
    columns the pivot and the Excel sheets read.

    `scope` is the DTE-filtered account map from dte_scope(): every compiled
    row whose canonical UserID is in it takes that account's Running Type
    (INT -> Intraday, POS -> Positional); every other row is Unclassified.

    `aliases` ({MTM key: All User key}, from load_user_aliases) covers the
    accounts the two files name differently — the MTM's XLDH142 is the
    sheet's CC04. It is tried FIRST, because the MTM id can itself exist in
    the sheet as a different (usually dropped, DLR ACC) row; the row's own id
    is still tried afterwards, so a stale alias can never cost a match that
    would otherwise have been made.

    Realized P&L is taken exactly as the MTM reports it — `AdjRealized` is
    the same number, kept under that name because the pivot, the raw sheet
    all read it."""
    comp = (compiled_source.copy() if isinstance(compiled_source, pd.DataFrame)
            else read_compiled(compiled_source, name))
    comp["UserID"] = comp["UserID"].astype(str).str.strip()

    for col in ["Realized P&L", "ALLOCATION", "MAX LOSS", "Unrealized P&L"]:
        if col in comp.columns:
            comp[col] = pd.to_numeric(comp[col], errors="coerce").fillna(0)
        else:
            comp[col] = 0
    if "SL HIT/NOT" not in comp.columns:
        comp["SL HIT/NOT"] = 0
    if "Alias" not in comp.columns:
        comp["Alias"] = ""

    # Matching is on the canonical user id ALONE: `userId` is unique in the
    # All User sheet, so the server adds nothing but a chance to mismatch
    # (the two files label servers independently).
    scope = scope or {}
    aliases = aliases or {}

    def lookup(key):
        alias = aliases.get(key)
        if alias is not None:
            hit = scope.get(alias)
            if hit is not None:
                return hit
        return scope.get(key)

    keys = comp["UserID"].map(_user_key)
    hits = [lookup(k) for k in keys]

    comp["Type"] = [h["type"] if h else UNCLASSIFIED for h in hits]
    comp["Running Type"] = [h["running_type"] if h else "" for h in hits]
    comp["Running Days"] = [h["running_days"] if h else "" for h in hits]
    # the All User id the row matched — differs from UserID only for aliased
    # accounts, so a glance at this column shows every alias that fired
    comp["All User ID"] = [
        (h["user_id"] if h and _user_key(h["user_id"]) != k else "")
        for h, k in zip(hits, keys)
    ]
    # reference only — MAX LOSS for every calculation is the compiled column
    comp["AllUser MaxLoss"] = [
        (h["max_loss"] if h and h["max_loss"] is not None else None) for h in hits
    ]

    comp["AdjRealized"] = comp["Realized P&L"]

    # MTM is ALWAYS computed — the compiled sheet ships an MTM column, but it
    # is only populated where Unrealized P&L is non-zero and left at 0 for
    # every other account (577 of 633 rows in the 30-07-2026 file), which
    # understates the book by the whole realized figure of those rows. The
    # column is overwritten rather than read.
    comp["MTM"] = comp["Realized P&L"] + comp["Unrealized P&L"]

    # Per-user return (MTM / Allocation) — the basis for MTM %, so every
    # percentage in the report sits on the same number as the money column
    # beside it.
    alloc = comp["ALLOCATION"]
    comp["UserReturn"] = (comp["MTM"] / alloc).where(alloc != 0, 0.0)
    return comp


def mtm_column_mismatch(compiled_source, name=None):
    """Data-quality probe on the uploaded sheet's own MTM column, so the fact
    that it is ignored and recomputed is visible rather than silent.

    Returns a dict describing what was actually MEASURED — deliberately no
    fixed narrative about the cause. Two different files have shown two
    different failure modes: one left MTM at 0 on 577 of 633 rows, another
    disagreed on just 2 rows. Reporting a hardcoded explanation would state a
    diagnosis the probe never made.

        rows      total rows examined
        mismatch  rows where the column != Realized + Unrealized
        blank     of those, rows where the column is 0 but the sum is not
                  (the "column never filled in" mode)
        differs   of those, rows carrying a genuinely different value
        shipped   sum of the column as uploaded
        computed  sum of Realized + Unrealized
        gap       computed - shipped, in rupees (the figure that matters:
                  a crore-scale display rounds sub-50,000 gaps away entirely)
        users     up to 10 affected UserIDs, for a spot check
    """
    df = (compiled_source if isinstance(compiled_source, pd.DataFrame)
          else read_compiled(compiled_source, name))
    empty = {"rows": len(df), "mismatch": 0, "blank": 0, "differs": 0,
             "shipped": 0.0, "computed": 0.0, "gap": 0.0, "users": []}
    if "MTM" not in df.columns:
        return empty
    shipped = pd.to_numeric(df["MTM"], errors="coerce").fillna(0)
    computed = (pd.to_numeric(df.get("Realized P&L"), errors="coerce").fillna(0)
                + pd.to_numeric(df.get("Unrealized P&L"), errors="coerce").fillna(0))
    bad = shipped.round(2) != computed.round(2)
    users = []
    if "UserID" in df.columns:
        users = [str(u) for u in df.loc[bad, "UserID"].head(10)]
    return {
        "rows": len(df),
        "mismatch": int(bad.sum()),
        "blank": int((bad & (shipped == 0)).sum()),
        "differs": int((bad & (shipped != 0)).sum()),
        "shipped": float(shipped.sum()),
        "computed": float(computed.sum()),
        "gap": float(computed.sum() - shipped.sum()),
        "users": users,
    }


def is_classified(comp):
    """True when at least one account carries an Int / Pos+Int type — i.e. the
    All User sheet matched something. False leaves the caller free to hide the
    per-type tiles."""
    return bool((comp["Type"] != UNCLASSIFIED).any())


def classified_only(comp):
    """The accounts the DTE scope covers — the per-algo blocks are built from
    these; the rest form the Unclassified section at the bottom."""
    return comp[comp["Type"] != UNCLASSIFIED]


def comp_stats(comp):
    n_pos = int((comp["Type"] == "Positional").sum())
    n_int = int((comp["Type"] == "Intraday").sum())
    n_unc = int((comp["Type"] == UNCLASSIFIED).sum())
    if not is_classified(comp):
        # nothing matched — the caller hides the per-type tiles
        return {"accounts": len(comp), "positional": None,
                "intraday": None, "unclassified": n_unc}
    return {
        "accounts": len(comp),
        "positional": n_pos,
        "intraday": n_int,
        "unclassified": n_unc,
    }


def infer_report_date(comp):
    """Best-effort report date (dd-mm-YYYY) from the compiled sheet's Date column."""
    if "Date" in comp.columns:
        for value in comp["Date"]:
            if pd.isna(value):
                continue
            if hasattr(value, "strftime"):
                return value.strftime("%d-%m-%Y")
            text = str(value).strip()
            if text:
                return text.split(" ")[0]
    return ""


# ====================== AGGREGATION ======================
def aggregate(sub: pd.DataFrame, by_algo=False):
    """One row per SERVER inside the block. `by_algo=True` groups by
    (ALGO, SERVER) instead — needed for the Unclassified section, which is NOT
    nested under one algo: grouping it by server alone would collapse two
    algos on the same server into a single row whose ALGO is whichever came
    first."""
    rows = []
    keys = ["ALGO", "SERVER"] if by_algo else ["SERVER"]
    for key, g in sub.groupby(keys, sort=True):
        server = key[-1] if isinstance(key, tuple) else key
        realized = g["AdjRealized"].sum()
        unreal   = g["Unrealized P&L"].sum()
        mtm      = g["MTM"].sum()
        alloc    = g["ALLOCATION"].sum()
        # the server's account detail — the deepest drill level of the pivot
        # (worst MTM first)
        user_rows = [
            {
                "UserID": rec["UserID"],
                "Alias": "" if pd.isna(rec["Alias"]) else str(rec["Alias"]),
                "SLHit": int(rec["SL HIT/NOT"] == 1),
                "MaxLoss": float(rec["MAX LOSS"]),
                "Allocation": float(rec["ALLOCATION"]),
                "Realized": float(rec["AdjRealized"]),
                "Unrealized": float(rec["Unrealized P&L"]),
                "MTM": float(rec["MTM"]),
                "Return": float(rec["UserReturn"]),
            }
            for _, rec in g.iterrows()
        ]
        user_rows.sort(key=lambda u: u["MTM"])
        rows.append({
            "ALGO"      : g["ALGO"].iloc[0],
            "SERVER"    : server,
            "Users"     : len(g),
            "SLHit"     : int((g["SL HIT/NOT"] == 1).sum()),
            "MaxLoss"   : g["MAX LOSS"].sum(),
            "Allocation": alloc,
            "Realized"  : realized,
            "Unrealized": unreal,
            "MTM"       : mtm,
            # the percentage follows the MTM column beside it
            "Return"    : (mtm / alloc) if alloc else 0.0,
            "UserRows"  : user_rows,
        })
    if by_algo:
        rows.sort(key=lambda x: (float(x["ALGO"]) if str(x["ALGO"]).replace(".", "", 1).isdigit()
                                 else float("inf"), str(x["SERVER"])))
    else:
        rows.sort(key=lambda x: str(x["SERVER"]))
    return rows


# ====================== STYLES ======================
# Display labels. Realized, Unrealized and MTM are all shown, so
# MTM = Realized P&L + Unrealized P&L can be verified at a glance, and MTM %
# is MTM / ALLOCATION — the percentage sits on the column beside it.
HEADERS = ["ALGO", "SERVER", "No. of Users", "No. of SL Hit Users", "MAX LOSS",
           "ALLOCATION", "Realized P&L", "Unrealized P&L", "MTM", "MTM %"]
KEYS    = ["ALGO", "SERVER", "Users", "SLHit", "MaxLoss",
           "Allocation", "Realized", "Unrealized", "MTM", "Return"]
NCOLS   = 1 + len(HEADERS)   # col A = type label, cols B-K = data

# Column letters (A=1, B=2, ..., K=11)
# B=ALGO  C=SERVER  D=Users  E=SLHit  F=MaxLoss
# G=Allocation  H=Realized  I=Unrealized  J=MTM  K=Return
KEY_COL   = {"Users": "D", "SLHit": "E", "MaxLoss": "F",
             "Allocation": "G", "Realized": "H", "Unrealized": "I", "MTM": "J"}
MONEY_KEYS = {"MaxLoss", "Allocation", "Realized", "Unrealized", "MTM"}

# --- Border sides ---
_thin = Side(style="thin",   color="C0C0C0")
_med  = Side(style="medium", color="595959")

def _bdr(l=None, r=None, t=None, b=None):
    return Border(left=l or _thin, right=r or _thin, top=t or _thin, bottom=b or _thin)

inner_bdr    = _bdr()
sub_bdr      = _bdr(t=_med, b=_med)     # medium top + bottom — sub-total reads as separator bar
banner_bdr   = _bdr(t=_med)             # top edge of each algo banner
algo_tot_bdr = _bdr(t=_med, b=_med)    # algo total: medium top + bottom
grand_bdr    = Border(left=_med, right=_med, top=_med, bottom=_med)

# --- Alignments ---
center      = Alignment(horizontal="center", vertical="center")
center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
vertical_a  = Alignment(horizontal="center", vertical="center", text_rotation=90)
left_wrap   = Alignment(horizontal="left", vertical="center", wrap_text=True)

# --- Fills ---
# Only title + header are dark. Every algo cycles through its own light palette pair.
title_fill  = PatternFill("solid", fgColor="1E293B")  # Deep Slate Blue/Gray  (title bar)
header_fill = PatternFill("solid", fgColor="64748B")  # Muted Cool Gray (column headers)
grand_fill  = PatternFill("solid", fgColor="D1C9E1")  # light lavender (grand total)

# Per-algo rotating palette — (data_hex, accent_hex)
#   data_hex   : ~10-12% saturation, used for ALL data rows in the algo
#   accent_hex : slightly more saturated, used for the banner, col-A label, algo-total row
#   sub-total  : derived = accent lightened 60% toward white, so it sits between the
#                data rows (lightest) and the accent, giving it its own distinct shade
ALGO_PALETTE = [
    ("FCE8E6", "E8A898"),  # blush pink
    ("E8F2E4", "98C490"),  # sage green
    ("EAE8F8", "9898D8"),  # periwinkle
    ("E4F0F8", "80BCDC"),  # sky blue
    ("F8F4E4", "C8B868"),  # warm cream / amber
    ("E8F6F2", "78BCA8"),  # seafoam / mint
    ("F8E8EE", "D898B0"),  # rose pink
    ("EEE8F8", "B090D0"),  # soft lilac
    ("E4F6F8", "68B0B8"),  # pale teal
]

def _lighten(hex_color, factor):
    """Lighten a hex colour by blending it `factor` of the way toward white."""
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    r = round(r + (255 - r) * factor)
    g = round(g + (255 - g) * factor)
    b = round(b + (255 - b) * factor)
    return f"{r:02X}{g:02X}{b:02X}"

def _algo_fills(i):
    data_hex, accent_hex = ALGO_PALETTE[i % len(ALGO_PALETTE)]
    sub_hex = _lighten(accent_hex, 0.60)   # sub-total: 60% lighter version of the accent
    return (PatternFill("solid", fgColor=data_hex),
            PatternFill("solid", fgColor=accent_hex),
            PatternFill("solid", fgColor=sub_hex))

white_bold = Font(bold=True, color="FFFFFF")
bold       = Font(bold=True)


# ====================== ROW WRITERS ======================
def _base(cell, fill, bdr, is_bold=False):
    cell.fill      = fill
    cell.border    = bdr
    cell.alignment = center
    if is_bold:
        cell.font = bold


def _algo_val(v):
    """Display algo number as integer when it is a whole float."""
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def write_data_row(ws, row, rowdict, fill, is_bold=False, bdr=None):
    b = bdr or inner_bdr
    for j, key in enumerate(KEYS):
        col  = j + 2
        cell = ws.cell(row, col)
        _base(cell, fill, b, is_bold)
        if key == "Return":
            # MTM % = MTM (col J) / ALLOCATION (col G) — the live formula keeps
            # the percentage tied to the MTM column when the sheet is edited
            cell.value         = f"=IF(G{row}=0,0,J{row}/G{row})"
            cell.number_format = "0.00"
        elif key == "ALGO":
            cell.value = _algo_val(rowdict[key])
        elif key in MONEY_KEYS:
            cell.value         = round(float(rowdict[key]))
            cell.number_format = INDIAN_XLSX_FMT
        else:
            cell.value = rowdict[key]


def write_subtotal_row(ws, row, data_start, data_end, fill, n_servers, is_bold=True, bdr=None):
    """Sub-total row: SUM formulas for all aggregated columns."""
    b = bdr or sub_bdr
    for j, key in enumerate(KEYS):
        col  = j + 2
        cell = ws.cell(row, col)
        _base(cell, fill, b, is_bold)
        if key == "ALGO":
            cell.value = "Sub-Total"
        elif key == "SERVER":
            cell.value = n_servers
        elif key in KEY_COL:
            cl             = KEY_COL[key]
            cell.value     = f"=SUM({cl}{data_start}:{cl}{data_end})"
            if key in MONEY_KEYS:
                cell.number_format = INDIAN_XLSX_FMT
        elif key == "Return":
            # MTM % = MTM (col J) / ALLOCATION (col G) — the live formula keeps
            # the percentage tied to the MTM column when the sheet is edited
            cell.value         = f"=IF(G{row}=0,0,J{row}/G{row})"
            cell.number_format = "0.00"


def write_total_row(ws, row, label, n_servers, ref_rows, fill, is_bold=True, bdr=None):
    """Algo-total or Grand-total: reference (sum) the given ref_rows."""
    b = bdr or algo_tot_bdr
    for j, key in enumerate(KEYS):
        col  = j + 2
        cell = ws.cell(row, col)
        _base(cell, fill, b, is_bold)
        if key == "ALGO":
            cell.value = label
        elif key == "SERVER":
            cell.value = n_servers
        elif key in KEY_COL:
            cl         = KEY_COL[key]
            formula    = "+".join(f"{cl}{rr}" for rr in ref_rows)
            cell.value = f"={formula}"
            if key in MONEY_KEYS:
                cell.number_format = INDIAN_XLSX_FMT
        elif key == "Return":
            # MTM % = MTM (col J) / ALLOCATION (col G) — the live formula keeps
            # the percentage tied to the MTM column when the sheet is edited
            cell.value         = f"=IF(G{row}=0,0,J{row}/G{row})"
            cell.number_format = "0.00"


def apply_outer_border(ws, start_row, end_row, start_col=1, end_col=NCOLS):
    """Overlay a medium border on the outer perimeter of a cell range.
    Preserves all inner borders that are already set."""
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row, col)
            b    = cell.border
            cell.border = Border(
                left   = _med if col == start_col else (b.left   or _thin),
                right  = _med if col == end_col   else (b.right  or _thin),
                top    = _med if row == start_row else (b.top    or _thin),
                bottom = _med if row == end_row   else (b.bottom or _thin),
            )


# ====================== PIVOT ROWS (for on-screen display) ======================
def pivot_rows(comp):
    """The pivot view as a flat list of dicts — same categorisation as the
    workbook's Summary sheet: per algo, an Int block (intraday accounts)
    and a Pos+Int block (positional accounts), each grouped by server. A
    server running both account types appears in both blocks, split by type.
    Each dict carries a `kind`: data / subtotal / algototal / grandtotal."""

    def block_row(kind, section, algo_label, server, block):
        realized = block["AdjRealized"].sum()
        unreal   = block["Unrealized P&L"].sum()
        mtm      = block["MTM"].sum()
        alloc    = block["ALLOCATION"].sum()
        return {
            "kind": kind, "Section": section,
            "ALGO": algo_label, "SERVER": server,
            "Users": len(block), "SLHit": int((block["SL HIT/NOT"] == 1).sum()),
            "MaxLoss": round(float(block["MAX LOSS"].sum())),
            "Allocation": round(float(alloc)),
            "Realized": round(float(realized)),
            "Unrealized": round(float(unreal)),
            "MTM": round(float(mtm)),
            "Return": round(float(mtm / alloc), 2) if alloc else 0.0,
        }

    def data_rows(sub, display_label, algo_val, by_algo=False):
        rows = []
        for rowdict in aggregate(sub, by_algo=by_algo):
            rows.append({
                "kind": "data", "Section": display_label, "algo_val": algo_val,
                "ALGO": _algo_val(rowdict["ALGO"]), "SERVER": rowdict["SERVER"],
                "Users": rowdict["Users"], "SLHit": rowdict["SLHit"],
                "MaxLoss": round(float(rowdict["MaxLoss"])),
                "Allocation": round(float(rowdict["Allocation"])),
                "Realized": round(float(rowdict["Realized"])),
                "Unrealized": round(float(rowdict["Unrealized"])),
                "MTM": round(float(rowdict["MTM"])),
                "Return": round(float(rowdict["Return"]), 2),
                "user_rows": [
                    {"UserID": u["UserID"], "Alias": u["Alias"], "SLHit": u["SLHit"],
                     "MaxLoss": round(u["MaxLoss"]), "Allocation": round(u["Allocation"]),
                     "Realized": round(u["Realized"]),
                     "Unrealized": round(u["Unrealized"]),
                     "MTM": round(u["MTM"]), "Return": round(u["Return"], 2)}
                    for u in rowdict["UserRows"]
                ],
            })
        return rows

    out = []
    # the per-algo blocks cover the DTE-scoped accounts only; whatever the
    # scope did not classify is reported once, at the bottom
    classified = classified_only(comp)
    unclassified = comp[comp["Type"] == UNCLASSIFIED]

    algos = sorted(classified["ALGO"].dropna().unique(), key=lambda x: float(x))
    total_n_servers = 0
    for algo in algos:
        algo_df    = classified[classified["ALGO"] == algo]
        algo_label = f"Algo {int(algo) if float(algo).is_integer() else algo}"
        algo_val   = _algo_val(algo)
        for ttype, display_label in [("Intraday", "Int"), ("Positional", "Pos+Int")]:
            sub = algo_df[algo_df["Type"] == ttype]
            if sub.empty:
                continue
            out.extend(data_rows(sub, display_label, algo_val))
            sub_row = block_row("subtotal", display_label, "Sub-Total",
                                sub["SERVER"].nunique(), sub)
            sub_row["algo_val"] = algo_val
            out.append(sub_row)
        algo_n = algo_df["SERVER"].nunique()
        total_row = block_row("algototal", "", f"{algo_label} Total", algo_n, algo_df)
        total_row["algo_val"] = algo_val
        out.append(total_row)
        total_n_servers += algo_n

    # ---- Unclassified section (own bucket, so the HTML pivot renders it as
    # one more collapsible group without any structural special-casing)
    if not unclassified.empty:
        out.extend(data_rows(unclassified, UNCLASSIFIED, UNCLASSIFIED, by_algo=True))
        sub_row = block_row("subtotal", UNCLASSIFIED, "Sub-Total",
                            unclassified["SERVER"].nunique(), unclassified)
        sub_row["algo_val"] = UNCLASSIFIED
        out.append(sub_row)
        unc_n = unclassified["SERVER"].nunique()
        total_row = block_row("algototal", "", f"{UNCLASSIFIED} Total",
                              unc_n, unclassified)
        total_row["algo_val"] = UNCLASSIFIED
        out.append(total_row)
        total_n_servers += unc_n

    grand_row = block_row("grandtotal", "", "Grand Total", total_n_servers, comp)
    grand_row["algo_val"] = ""
    out.append(grand_row)
    return out


# ====================== SLIPPAGE (REALIZED LOSS % vs MAX-LOSS %) ======================
SLIP_THRESHOLD = 0.1   # of allocation — same ratio units as the values themselves
_SLIP_EPS = 1e-9       # float guard so a difference of exactly 0.1 counts


def no_sl_accounts(comp):
    """Accounts trading with NO stop-loss configured — allocation > 0 but
    MAX LOSS <= 0 (or blank, which prepare_comp reads as 0).

    They are excluded from the slippage analysis entirely: ML % would be 0,
    so ANY realized loss past 10% of allocation would be reported as
    "slippage" against a limit that was never set. That is a configuration
    gap, not a stop that slipped — it is listed in its own "no_sl_Acc" sheet
    so it can be fixed at source."""
    sel = comp[(comp["ALLOCATION"] > 0) & (comp["MAX LOSS"] <= 0)]
    return [
        {
            "ALGO": r["ALGO"], "SERVER": r["SERVER"], "UserID": r["UserID"],
            "Alias": "" if pd.isna(r["Alias"]) else r["Alias"],
            "Allocation": float(r["ALLOCATION"]),
            "MaxLoss": float(r["MAX LOSS"]),
            "Realized": float(r["Realized P&L"]),
            "Unrealized": float(r["Unrealized P&L"]),
            "MTM": float(r["MTM"]),
            "RunningType": r.get("Running Type", ""),
            "RunningDays": r.get("Running Days", ""),
        }
        for _, r in sel.iterrows()
    ]


def slippage_eligible(comp):
    """The accounts the slippage analysis actually covers: a real allocation
    AND a configured max loss. Used for the rows and for the account counts,
    so the "Accounts" column never includes accounts that could not have been
    judged in the first place."""
    return comp[(comp["ALLOCATION"] > 0) & (comp["MAX LOSS"] > 0)]


def slippage_rows(comp):
    """One row per SLIPPAGE account, evaluated over the ELIGIBLE accounts —
    allocation > 0 and a configured MAX LOSS (see no_sl_accounts for the
    accounts excluded and why). Both limits are plain RATIOS of the account's
    allocation (positive numbers = loss) — the same unit convention as the
    pivot's MTM %, NOT multiplied by 100:

        ML %          = MAX LOSS / ALLOCATION
        Realized ML % = -compiled Realized P&L / ALLOCATION

    An account HAS SLIPPAGE only when its realized loss overshoots the
    configured max-loss by at least 0.1:

        Realized ML % - ML % >= 0.1

    (ML% 1.00 -> Realized 1.09 is NOT slippage; 1.10 is. A profit or a loss
    inside the limit is never slippage.) Sorted worst-first (largest
    Realized ML %)."""
    rows = []
    for _, r in slippage_eligible(comp).iterrows():
        alloc = float(r["ALLOCATION"])
        ml_pct = float(r["MAX LOSS"]) / alloc
        realized_ml_pct = -float(r["Realized P&L"]) / alloc
        if realized_ml_pct - ml_pct < SLIP_THRESHOLD - _SLIP_EPS:
            continue
        rows.append({
            "ALGO": r["ALGO"],
            "SERVER": r["SERVER"],
            "UserID": r["UserID"],
            "Alias": "" if pd.isna(r["Alias"]) else r["Alias"],
            "Allocation": alloc,
            "MaxLoss": float(r["MAX LOSS"]),
            "Realized": float(r["Realized P&L"]),
            "MLPct": ml_pct,
            "RealizedMLPct": realized_ml_pct,
            "DiffPct": realized_ml_pct - ml_pct,
        })
    rows.sort(key=lambda x: -x["RealizedMLPct"])
    return rows


def slippage_summary(rows, comp):
    """Per algo + overall: total accounts, slippage accounts, and the
    AVERAGE SLIPPAGE of the algo = the mean Realized ML % over the algo's
    slippage accounts (None when the algo has none).

    The account counts are over the ELIGIBLE accounts only, the same set
    slippage_rows judges — counting accounts that carry no max loss (or no
    allocation) would report a denominator the analysis never looked at."""
    def block(sub, n_accounts):
        return {
            "accounts": n_accounts,
            "slipped": len(sub),
            "avg_slippage": (sum(r["RealizedMLPct"] for r in sub) / len(sub)) if sub else None,
        }

    eligible = slippage_eligible(comp)
    algos = sorted(eligible["ALGO"].dropna().unique(), key=float)
    per_algo = [{"ALGO": _algo_val(algo),
                 **block([r for r in rows if r["ALGO"] == algo],
                         int((eligible["ALGO"] == algo).sum()))}
                for algo in algos]
    return per_algo, {"ALGO": "Overall", **block(rows, len(eligible))}


def major_slippages(rows, per_algo):
    """Major slippage = a slippage account whose Realized ML % is GREATER
    than its algo's average slippage.

    An algo with exactly ONE slippage account is a special case: that account
    IS the average, so a strict `>` test can never fire and the algo could
    never report a major however badly the account overshot. A lone slippage
    account is the worst in its algo by definition, so it is reported as
    major. (With two accounts the mean sits between them and exactly one
    qualifies; the test only becomes discriminating from three up.)

    Each row carries the algo average for display; the worst-first order of
    `rows` is preserved."""
    avg_by_algo = {s["ALGO"]: s["avg_slippage"] for s in per_algo}
    n_by_algo = {s["ALGO"]: s["slipped"] for s in per_algo}
    out = []
    for r in rows:
        algo = _algo_val(r["ALGO"])
        avg = avg_by_algo.get(algo)
        if avg is None:
            continue
        if n_by_algo.get(algo) == 1 or r["RealizedMLPct"] > avg:
            out.append({**r, "AlgoAvgSlippage": avg})
    return out


UNCLASSIFIED_HEADERS = ["UserID", "Alias", "ALGO", "SERVER",
                        "Running Type", "Running Days"]

# what the unclassified sheet prints when the All User sheet has nothing for
# an account — dropped on upload (DLR ACC / NOT RUNNING) or absent entirely
NOT_FOUND = "Not Found"


def _write_unclassified_sheet(wb, comp, all_user_ref=None, suffix=""):
    """Sheet "unclassified", placed FIRST in the workbook: the accounts the DTE
    scope did not cover.

    ALGO and SERVER are the compiled MTM's own values — the account as it
    actually traded. Running Type / Running Days are looked up in the All User
    sheet AFTER the DLR ACC / NOT RUNNING drop:

        INT / 0DTE  (in a 4DTE run) running, but not on this DTE
        Not Found                   dropped on upload, or never in the sheet

    A dropped row is gone, so it reads "Not Found" exactly like an id the
    sheet never carried — the report does not resurrect a row it dropped."""
    ref = all_user_ref or {}
    rows = comp[comp["Type"] == UNCLASSIFIED]

    ws = wb.create_sheet("unclassified" + suffix)
    for j, h in enumerate(UNCLASSIFIED_HEADERS, start=1):
        cell = ws.cell(1, j, h)
        cell.fill = header_fill
        cell.font = white_bold
        cell.alignment = center_wrap
        cell.border = inner_bdr

    r = 2
    for _, rec in rows.iterrows():
        # dropped on upload, or never in the sheet — either way the All User
        # sheet has nothing to say about this account
        hit = ref.get(_user_key(rec["UserID"]))
        values = [
            rec["UserID"],
            "" if pd.isna(rec["Alias"]) else str(rec["Alias"]),
            _algo_val(rec["ALGO"]),
            rec["SERVER"],
            hit["running_type"] if hit else NOT_FOUND,
            hit["running_days"] if hit else NOT_FOUND,
        ]
        for j, v in enumerate(values, start=1):
            cell = ws.cell(r, j, v)
            cell.border = inner_bdr
            if j >= 3:
                cell.alignment = center
        r += 1
    if rows.empty:
        ws.merge_cells(start_row=r, start_column=1, end_row=r,
                       end_column=len(UNCLASSIFIED_HEADERS))
        nc = ws.cell(r, 1, "-- Every account was classified --")
        nc.alignment = center
        nc.font = Font(italic=True, color="595959")
        r += 1

    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1,
                   end_column=len(UNCLASSIFIED_HEADERS))
    note = ws.cell(
        r + 1, 1,
        "*Accounts in the User MTM that the selected DTE scope does not cover. ALGO and "
        "SERVER are as the MTM reports them. Running Type / Running Days are looked up in "
        "the All User sheet AFTER the DLR ACC / NOT RUNNING rows are dropped: a running "
        "pair (e.g. INT / 0DTE in a 4DTE run) means the account runs on a different DTE, "
        "and \"Not Found\" means it was dropped on upload or is not in the sheet at all."
    )
    note.font = Font(italic=True, color="595959", size=9)
    note.alignment = left_wrap
    ws.row_dimensions[r + 1].height = 42

    ws.freeze_panes = "A2"
    for j, w in enumerate([20, 30, 8, 10, 15, 15], start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # move to the FRONT of the workbook, after any earlier unclassified sheet
    # (a second User MTM adds its own, which must not jump ahead of the first)
    target = sum(1 for s in wb.worksheets
                 if s.title.startswith("unclassified") and s is not ws)
    wb.move_sheet(ws, offset=target - wb.worksheets.index(ws))


NO_SL_HEADERS = ["ALGO", "SERVER", "UserID", "Alias", "Running Type", "Running Days",
                 "ALLOCATION", "MAX LOSS", "Realized P&L", "Unrealized P&L", "MTM"]


def _write_no_sl_sheet(wb, comp, report_date, suffix=""):
    """Sheet "no_sl_Acc": every account trading with allocation but NO max
    loss configured. These are excluded from the slippage analysis (ML % would
    be 0, so any loss past the threshold would read as slippage against a
    limit that was never set) and listed here so the gap can be closed."""
    rows = no_sl_accounts(comp)
    ws = wb.create_sheet("no_sl_Acc" + suffix)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(NO_SL_HEADERS))
    tc = ws.cell(1, 1, f"Accounts with NO Max Loss configured  ({report_date})")
    tc.font = Font(bold=True, italic=True, color="FFFFFF", size=13)
    tc.fill = title_fill
    tc.alignment = center
    ws.row_dimensions[1].height = 22

    for j, h in enumerate(NO_SL_HEADERS, start=1):
        cell = ws.cell(2, j, h)
        cell.fill = header_fill
        cell.font = white_bold
        cell.alignment = center_wrap
        cell.border = inner_bdr

    money = {"ALLOCATION", "MAX LOSS", "Realized P&L", "Unrealized P&L", "MTM"}
    r = 3
    for row in rows:
        values = [_algo_val(row["ALGO"]), row["SERVER"], row["UserID"], row["Alias"],
                  row["RunningType"], row["RunningDays"],
                  round(row["Allocation"]), round(row["MaxLoss"]),
                  round(row["Realized"]), round(row["Unrealized"]), round(row["MTM"])]
        for j, v in enumerate(values, start=1):
            cell = ws.cell(r, j, v)
            cell.border = inner_bdr
            cell.alignment = center
            if NO_SL_HEADERS[j - 1] in money:
                cell.number_format = INDIAN_XLSX_FMT
        r += 1
    if not rows:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(NO_SL_HEADERS))
        nc = ws.cell(r, 1, "-- Every account has a max loss configured --")
        nc.alignment = center
        nc.font = Font(italic=True, color="595959")
        r += 1

    note = ws.cell(
        r + 1, 1,
        "*These accounts have an allocation but MAX LOSS = 0, so no stop-loss limit is set. "
        "They are EXCLUDED from the Max SL Slippage Analysis — with ML % = 0 any loss past "
        "the threshold would be reported as slippage against a limit that never existed."
    )
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1,
                   end_column=len(NO_SL_HEADERS))
    note.font = Font(italic=True, color="595959", size=9)
    note.alignment = left_wrap
    ws.row_dimensions[r + 1].height = 32

    ws.freeze_panes = "A3"
    for j, w in enumerate([7, 10, 20, 24, 13, 13, 14, 12, 15, 15, 15], start=1):
        ws.column_dimensions[get_column_letter(j)].width = w


SLIP_SUMMARY_HEADERS = ["ALGO", "Accounts", "Slippage Accounts", "Avg Slippage %"]
SLIP_DETAIL_HEADERS = ["ALGO", "SERVER", "UserID", "Alias", "ALLOCATION", "MAX LOSS",
                       "Realized P&L", "ML %", "Realized ML %", "Diff %"]
SLIP_MAJOR_HEADERS = SLIP_DETAIL_HEADERS + ["Algo Avg Slippage %"]


def _write_slippage_sheet(wb, comp, report_date, suffix=""):
    """Sheet "Slippage": the per-algo summary on top (accounts / slippage
    accounts / avg slippage %), then every slippage account worst-first with
    the MAJOR rows (Realized ML % above the algo average) highlighted."""
    rows = slippage_rows(comp)
    per_algo, overall = slippage_summary(rows, comp)
    # the highlight must use the SAME rule as the report's Major table — ask
    # major_slippages rather than re-implementing the test here, or the sheet
    # and the report can disagree about what counts as major
    major_keys = {(m["UserID"], m["SERVER"]) for m in major_slippages(rows, per_algo)}

    ws4 = wb.create_sheet("Slippage" + suffix)
    n_cols = max(len(SLIP_SUMMARY_HEADERS), len(SLIP_DETAIL_HEADERS))

    ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    tc = ws4.cell(1, 1, f"Max SL Slippage Analysis — Realized Loss % vs Max Loss % of Allocation  ({report_date})")
    tc.font = Font(bold=True, italic=True, color="FFFFFF", size=13)
    tc.fill = title_fill
    tc.alignment = center
    ws4.row_dimensions[1].height = 22

    def header_row(r, headers):
        for j, h in enumerate(headers, start=1):
            cell = ws4.cell(r, j, h)
            cell.fill = header_fill
            cell.font = white_bold
            cell.alignment = center_wrap
            cell.border = inner_bdr

    r = 2
    header_row(r, SLIP_SUMMARY_HEADERS)
    r += 1
    grand_row_fill = PatternFill("solid", fgColor="D1C9E1")
    for srow in per_algo + [overall]:
        is_overall = srow["ALGO"] == "Overall"
        values = [srow["ALGO"], srow["accounts"], srow["slipped"],
                  round(srow["avg_slippage"], 2) if srow["avg_slippage"] is not None else "—"]
        for j, v in enumerate(values, start=1):
            cell = ws4.cell(r, j, v)
            cell.alignment = center
            cell.border = inner_bdr
            if is_overall:
                cell.fill = grand_row_fill
                cell.font = bold
            if j == 4 and v != "—":
                cell.number_format = "0.00"
        r += 1

    r += 1  # spacer
    detail_header = r
    if rows:
        header_row(r, SLIP_DETAIL_HEADERS)
        r += 1
        major_fill = PatternFill("solid", fgColor="FCE8E6")   # major slippage rows
        for row in rows:
            is_major = (row["UserID"], row["SERVER"]) in major_keys
            values = [_algo_val(row["ALGO"]), row["SERVER"], row["UserID"], row["Alias"],
                      round(row["Allocation"]), round(row["MaxLoss"]), round(row["Realized"]),
                      round(row["MLPct"], 2), round(row["RealizedMLPct"], 2),
                      round(row["DiffPct"], 2)]
            for j, v in enumerate(values, start=1):
                cell = ws4.cell(r, j, v)
                cell.alignment = center
                cell.border = inner_bdr
                if is_major:
                    cell.fill = major_fill
                if j in (5, 6, 7):
                    cell.number_format = INDIAN_XLSX_FMT
                elif j in (8, 9, 10):
                    cell.number_format = "0.00"
            r += 1
    else:
        ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
        nc = ws4.cell(r, 1, "-- No slippage today --")
        nc.alignment = center
        nc.font = Font(italic=True, color="595959")
        r += 1

    note_row = r + 1
    ws4.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=n_cols)
    nc = ws4.cell(
        note_row, 1,
        "*ML % = MAX LOSS / ALLOCATION; Realized ML % = |Realized P&L| / ALLOCATION "
        "— measured on the REALIZED loss, not MTM, because a max-loss stop is about what "
        "was actually booked. Plain ratios of allocation, same convention as MTM %. "
        "An account has slippage "
        "only when Realized ML % exceeds ML % by at least 0.1 (1.00 → 1.09 is not slippage; "
        "1.10 is). Avg Slippage = average Realized ML % of the algo's slippage accounts; "
        "highlighted rows are MAJOR slippages — accounts above their algo's average, or the "
        "lone slippage account of an algo (which is that algo's worst by definition). "
        "Accounts with no max loss configured are excluded — see the no_sl_Acc sheet."
    )
    nc.font = Font(italic=True, color="595959", size=9)
    nc.alignment = left_wrap
    ws4.row_dimensions[note_row].height = 40

    ws4.freeze_panes = f"A{detail_header + 1}"
    slip_widths = [7, 10, 20, 22, 13, 13, 14, 9, 13, 9]
    for j, w in enumerate(slip_widths, start=1):
        ws4.column_dimensions[get_column_letter(j)].width = w


# ====================== WORKBOOK ======================
def add_summary_sheets(wb, comp, report_date, suffix="", all_user_ref=None):
    """Append the Summary pivot + MTM Data +
    Slippage sheets to an existing openpyxl workbook (used to combine reports
    in one file). `suffix` (e.g. " 2") renames every sheet so a second MTM
    file's set can sit beside the first."""
    ws = wb.create_sheet("Summary" + suffix)

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
    tc = ws.cell(1, 1, f"Algo & Server wise MTM  —  Int / Pos+Int  ({report_date})")
    tc.font      = Font(bold=True, italic=True, color="FFFFFF", size=13)
    tc.fill      = title_fill
    tc.alignment = center

    # Header row
    HEADER_ROW = 2
    ws.cell(HEADER_ROW, 1, "Type")
    for j, h in enumerate(HEADERS):
        ws.cell(HEADER_ROW, j + 2, h)
    for c in range(1, NCOLS + 1):
        cell           = ws.cell(HEADER_ROW, c)
        cell.fill      = header_fill
        cell.font      = white_bold
        cell.alignment = center_wrap
        cell.border    = inner_bdr

    # ---- Build sheet ----
    r               = HEADER_ROW + 1
    algo_total_rows = []    # row numbers of each algo-total row (for grand total formula)
    total_n_servers = 0

    # per-algo blocks cover the DTE-scoped accounts; the rest get one section
    # of their own at the bottom (normally empty)
    classified   = classified_only(comp)
    unclassified = comp[comp["Type"] == UNCLASSIFIED]
    algos = sorted(classified["ALGO"].dropna().unique(), key=lambda x: float(x))

    def write_block(block_df, banner_label, sections, palette_index, by_algo=False):
        """One banner + its Int/Pos+Int (or Unclassified) sections + the block
        total row. Returns the block-total row number and its server count."""
        nonlocal r
        d_fill, a_fill, s_fill = _algo_fills(palette_index)

        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=NCOLS)
        bc           = ws.cell(r, 2, banner_label)
        bc.fill      = a_fill
        bc.font      = Font(bold=True, color="1F2D3D", size=11)
        bc.alignment = center
        for c in range(1, NCOLS + 1):
            cell        = ws.cell(r, c)
            cell.fill   = a_fill
            cell.border = banner_bdr
        ws.row_dimensions[r].height = 20
        r += 1

        section_sub_rows = []
        for ttype, display_label in sections:
            sub = block_df[block_df["Type"] == ttype]
            if sub.empty:
                continue

            rows        = aggregate(sub, by_algo=by_algo)
            block_start = r

            for rowdict in rows:
                write_data_row(ws, r, rowdict, d_fill)
                ws.row_dimensions[r].height = 18
                r += 1

            data_end = r - 1
            sub_row  = r
            # sub-total uses its own 60%-lighter accent shade so it reads distinct from data rows
            write_subtotal_row(ws, r, block_start, data_end, s_fill,
                               n_servers=len(rows))
            ws.row_dimensions[r].height = 18
            r += 1

            # Section label col A — merged, rotated, accent colour
            ws.merge_cells(start_row=block_start, start_column=1, end_row=sub_row, end_column=1)
            lc           = ws.cell(block_start, 1, display_label)
            lc.fill      = a_fill
            lc.font      = Font(bold=True, color="1F2D3D")
            lc.alignment = vertical_a
            for rr in range(block_start, sub_row + 1):
                ws.cell(rr, 1).fill   = a_fill
                ws.cell(rr, 1).border = inner_bdr

            # Outer medium border boxes the whole Int / Pos+Int section
            apply_outer_border(ws, block_start, sub_row)

            section_sub_rows.append(sub_row)

        # Block-level server count is DISTINCT across its sections: a server that
        # runs both types appears in each section but is counted once here.
        n_servers = block_df["SERVER"].nunique()

        # ---- Block Total — col A:B merged & centered (no orphan empty cell) ----
        write_total_row(ws, r, f"{banner_label} Total", n_servers,
                        section_sub_rows, a_fill)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        mc           = ws.cell(r, 1)          # top-left cell owns the merged range
        mc.value     = f"{banner_label} Total"
        mc.fill      = a_fill
        mc.font      = bold
        mc.alignment = center
        mc.border    = algo_tot_bdr
        ws.row_dimensions[r].height = 20
        apply_outer_border(ws, r, r)
        total_row = r
        r += 2   # block total, then a blank spacer row
        return total_row, n_servers

    for i, algo in enumerate(algos):
        algo_label = f"Algo {int(algo) if float(algo).is_integer() else algo}"
        total_row, n_servers = write_block(
            classified[classified["ALGO"] == algo], algo_label,
            [("Intraday", "Int"), ("Positional", "Pos+Int")], i)
        algo_total_rows.append(total_row)
        total_n_servers += n_servers

    # ---- Unclassified: accounts the DTE scope did not cover ----
    if not unclassified.empty:
        total_row, n_servers = write_block(
            unclassified, UNCLASSIFIED,
            [(UNCLASSIFIED, "Unc")], len(algos), by_algo=True)
        algo_total_rows.append(total_row)
        total_n_servers += n_servers

    # ---- Grand total ----
    write_total_row(ws, r, "Grand Total", total_n_servers, algo_total_rows,
                    grand_fill, bdr=grand_bdr)
    gc           = ws.cell(r, 1, "Total")
    gc.fill      = grand_fill
    gc.font      = bold
    gc.alignment = center
    gc.border    = grand_bdr

    # ---- P&L sign colouring: Realized (H) / Unrealized (I) / MTM (J) ----
    # profit green, loss red — conditional formatting so the live formulas
    # keep recolouring when the sheet is edited
    pnl_range = f"H{HEADER_ROW + 1}:J{r}"
    ws.conditional_formatting.add(pnl_range, CellIsRule(
        operator="greaterThan", formula=["0"], font=Font(color="15803D")))
    ws.conditional_formatting.add(pnl_range, CellIsRule(
        operator="lessThan", formula=["0"], font=Font(color="DC2626")))

    # ---- Footnote: how MTM is derived ----
    note_row = r + 2
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=NCOLS)
    nc = ws.cell(
        note_row, 1,
        "*MTM = Realized P&L + Unrealized P&L, computed per account from the uploaded User "
        "MTM — the sheet's own MTM column is not used, it is populated only where Unrealized "
        "is non-zero. MTM % = MTM / ALLOCATION. Every figure here aggregates the "
        "\"MTM Data\" sheet."
    )
    nc.font      = Font(italic=True, color="595959", size=9)
    nc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[note_row].height = 32

    # ---- Finish formatting ----
    ws.freeze_panes = f"A{HEADER_ROW + 1}"
    ws.column_dimensions["A"].width = 5
    widths = [7, 9, 12, 18, 13, 13, 15, 15, 15, 10]
    for j, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(j + 2)].width = w
    ws.row_dimensions[1].height = 22

    # the base data and slippage sheets cover EVERY account, classified or not
    _write_raw_sheet(wb, comp, suffix)
    _write_slippage_sheet(wb, comp, report_date, suffix)
    _write_no_sl_sheet(wb, comp, report_date, suffix)
    # written last, moved to the front of the workbook
    _write_unclassified_sheet(wb, comp, all_user_ref, suffix)


def build_workbook(comp, report_date, all_user_ref=None):
    """Build the full styled workbook (Summary pivot + MTM Data +
    MTM Data + Slippage) and return it."""
    wb = Workbook()
    wb.remove(wb.active)
    add_summary_sheets(wb, comp, report_date, all_user_ref=all_user_ref)
    return wb


# ====================== MTM DATA (per-user base sheet) ======================
# The single flat table every summary figure is computed from: the compiled
# User MTM's own columns, plus Running Type / Running Days from the All User
# sheet, plus MTM computed as Realized P&L + Unrealized P&L. Reading down this
# sheet reproduces the pivot exactly — which is the point of it existing.
MTM_DATA_COLS = [
    ("UserID",         "UserID"),
    ("Alias",          "Alias"),
    ("ALLOCATION",     "ALLOCATION"),
    ("MAX LOSS",       "MAX LOSS"),
    ("Total Orders",   "Total Orders"),
    ("Total Lots",     "Total Lots"),
    ("SERVER",         "SERVER"),
    ("ALGO",           "ALGO"),
    ("Running Type",   "Running Type"),
    ("Running Days",   "Running Days"),
    ("OPERATOR",       "OPERATOR"),
    ("EXPIRY",         "EXPIRY"),
    ("Date",           "Date"),
    ("Month",          "Month"),
    ("Day",            "Day"),
    ("INDEX",          "INDEX"),
    ("MTM",            "MTM"),
    ("Realized P&L",   "Realized P&L"),
    ("Unrealized P&L", "Unrealized P&L"),
    ("SL HIT/NOT",     "SL Hit"),
]
MTM_DATA_MONEY = {"ALLOCATION", "MAX LOSS", "MTM", "Realized P&L", "Unrealized P&L"}
MTM_DATA_COUNT = {"Total Orders", "Total Lots"}
MTM_DATA_CENTER = {"SERVER", "ALGO", "Running Type", "Running Days", "EXPIRY",
                   "Date", "Month", "Day", "INDEX", "SL HIT/NOT"}


def _write_raw_sheet(wb, comp, suffix=""):
    comp = comp.copy()
    comp["SL Hit"] = (comp["SL HIT/NOT"] == 1).astype(int)

    # a compiled MTM that does not carry one of the descriptive columns
    # (Total Orders, OPERATOR, INDEX, …) gets it blank rather than failing
    for _, src in MTM_DATA_COLS:
        if src not in comp.columns:
            comp[src] = ""

    raw = comp[[src for _, src in MTM_DATA_COLS]].copy()
    raw.columns = [hdr for hdr, _ in MTM_DATA_COLS]
    raw = raw.sort_values(["ALGO", "SERVER", "UserID"]).reset_index(drop=True)

    ws2 = wb.create_sheet("MTM Data" + suffix)
    for j, hdr in enumerate(raw.columns, start=1):
        cell           = ws2.cell(1, j, hdr)
        cell.fill      = header_fill
        cell.font      = white_bold
        cell.alignment = center_wrap
        cell.border    = inner_bdr

    for i, rec in enumerate(raw.itertuples(index=False), start=2):
        for j, (hdr, val) in enumerate(zip(raw.columns, rec), start=1):
            # a column the compiled sheet never carried reads NaN — leave the
            # cell empty rather than writing a value Excel cannot format
            if pd.isna(val):
                val = None
            cell        = ws2.cell(i, j, val)
            cell.border = inner_bdr
            if hdr in MTM_DATA_MONEY and isinstance(val, (int, float)):
                cell.value         = float(val)
                cell.number_format = INDIAN_XLSX_FMT
            elif hdr in MTM_DATA_COUNT and isinstance(val, (int, float)):
                cell.number_format = INDIAN_XLSX_FMT
            if hdr in MTM_DATA_CENTER:
                cell.alignment = center

    ws2.freeze_panes = "A2"
    raw_widths = [20, 24, 14, 14, 13, 12, 9, 7, 13, 13, 14, 12,
                  12, 10, 11, 14, 15, 15, 15, 11]
    for j, w in enumerate(raw_widths, start=1):
        ws2.column_dimensions[get_column_letter(j)].width = w


# ====================== CLI ======================
def ask_dte():
    """Ask which DTE the report covers — it decides which All User accounts
    are in scope (0DTE = every running account)."""
    while True:
        raw = input(f"  >> DTE ({' / '.join(DTE_OPTIONS)}):\n     DTE: ").strip().upper()
        if raw in DTE_RUNNING_DAYS:
            return raw
        print(f"     [!] Enter one of: {', '.join(DTE_OPTIONS)}\n")


def main():
    print("=" * 60)
    print("  SEGREGATION REPORT - INPUTS")
    print("=" * 60)
    report_date   = ask_date("Enter the report date (e.g. 09-07-2026):")
    dte           = ask_dte()
    compiled_path = ask_file("Path to the Updated Compiled User MTM file:", required=True)
    all_user_path = ask_file(
        f"Path to the All User Details file (tab \"{ALL_USER_SHEET}\"):", required=True)

    print("\nReading source files...")
    all_users = read_all_users(all_user_path)
    scope     = dte_scope(all_users, dte)
    aliases   = load_user_aliases()
    comp      = prepare_comp(compiled_path, scope, aliases=aliases)
    user_ref  = all_user_reference(all_users)
    if aliases:
        print(f"  Aliases loaded: {len(aliases)} from {ALIAS_FILE}")
        fired = int((comp["All User ID"].astype(str) != "").sum())
        print(f"  Aliases applied: {fired} compiled account(s) matched under "
              f"their All User id")

    in_scope, n_pos_scope, n_int_scope = scope_stats(all_users, dte)
    stats = comp_stats(comp)
    print(f"  All User -> {len(all_users)} running account(s) after the "
          f"DLR ACC / NOT RUNNING drop")
    print(f"  {dte} scope -> {in_scope} account(s)  |  POS: {n_pos_scope}  |  INT: {n_int_scope}")
    print(f"  Matched in the MTM -> Accounts: {stats['accounts']}  |  "
          f"Positional: {stats['positional']}  |  Intraday: {stats['intraday']}")
    if stats["unclassified"]:
        print(f"  [!] {stats['unclassified']} compiled account(s) are outside the "
              f"{dte} scope — reported in the Unclassified section")

    print("Writing raw per-user data sheet...")
    wb = build_workbook(comp, report_date, all_user_ref=user_ref)

    output_path = os.path.join(
        SCRIPT_DIR, f"Segregated_Int_Pos_MTM_{dte}_{report_date}.xlsx")
    wb.save(output_path)

    py_realized = comp["Realized P&L"].sum()
    py_mtm      = comp["MTM"].sum()
    print(f"\nDone. Report saved at:\n  {output_path}")
    print(f"  Grand Total -> Users: {stats['accounts']}  "
          f"Realized: {round(py_realized):,}  MTM: {round(py_mtm):,}")


if __name__ == "__main__":
    main()
