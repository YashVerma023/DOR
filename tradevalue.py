"""Trade value report.

Reads a compiled orderbook (CSV/Excel) plus the compiled user MTM summary
(CSV/Excel) and produces tradevalue.csv: trade value / lots per
(date, server, user, segment), with each user's ALLOCATION pulled from the
summary sheet (matched on user id).

Rules (same as the original report):
  - only F&O rows count: Exchange must be BFO or NFO (drops NSE/BSE/MCX cash
    rows like NIFTYBEES-EQ, whose quantities aren't lot multiples);
  - only COMPLETE orders count;
  - duplicate orders are dropped on (user_id, date, order_id, exchg_order_id),
    keeping the first occurrence (lowest SNO);
  - symbols are classified into NIFTY / BANKNIFTY / SENSEX segments
    (anything else is ignored);
  - lots = abs(quantity) / lot size, using the date-based lot-size history;
  - trade value = sum(avg price * quantity).

Usage:
    python tradevalue.py                                 # auto-pick files in this folder
    python tradevalue.py orderbook.csv -s user_mtm.xlsx  # explicit inputs
    python tradevalue.py -o report.csv                   # custom output name
    streamlit run app.py                                 # browser UI with file upload
"""

import argparse
import csv
import io
import logging
import re
import statistics
import sys
from collections import Counter, namedtuple
from statistics import NormalDist
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

logger = logging.getLogger(__name__)

SCRIPT_DIR = Path(__file__).resolve().parent
ORDERBOOK_PATTERN = "Compiled_Orderbook_*"
SUMMARY_PATTERN = "Compiled_User_MTM_*"
TABLE_SUFFIXES = (".csv", ".xlsx", ".xlsm", ".xls")
DEFAULT_OUTPUT = SCRIPT_DIR / "tradevalue.xlsx"

REPORT_HEADER = ["Date", "Server", "User ID", "Allocation", "Algo", "Type", "Segment",
                 "Order Count", "Quantity", "Lots", "Normalise", "Lots per Cr",
                 "Trade Value", "Outlier"]

SUMMARY_HEADER = ["Algo", "Type", "Total Users", "Median Lots per Cr", "Range",
                  "Below", "In", "Above"]

# Normalise base: allocation / 1,00,000 (stored value; 1,00,000 stored = 1 Cr
# in the x100 display convention). Every account normalises by it — 15,60,000
# -> 15.6, and the sub-1-Cr variations 80,000 / 60,000 / 40,000 / 20,000
# -> 0.8 / 0.6 / 0.4 / 0.2.
NORMALISE_BASE = Decimal(100000)

STRIKE_ALGO_HEADER = ["Algo", "Server", "Strikes Traded"]
STRIKE_CHAIN_HEADER = ["CE", "Strike", "PE"]
ORDER_SUMMARY_HEADER = ["Algo", "Type", "Total Users", "Total Orders", "Executed",
                        "Failed/Cancelled/Rejected", "Pending", "Hedge", "VAR"]
ORDER_SERVER_HEADER = ["Server", "Users", "Total Orders", "Executed",
                       "Failed/Cancelled/Rejected", "Pending", "Hedge", "VAR"]

# An order that is still live at end of day is NOT a failure — the real
# orderbook carries OPEN / OPEN_PENDING alongside CANCELLED and REJECTED, and
# lumping them together would report live orders as rejected.
PENDING_STATUSES = {"OPEN", "OPEN_PENDING", "TRIGGER_PENDING", "TRIGGER PENDING",
                    "AMO_SUBMITTED", "PENDING", "NEW"}

# The newer export writes "Filled" where the compiled orderbook wrote
# "COMPLETE", and both spellings of cancelled appear. Normalised on read so
# every downstream test can keep comparing against one vocabulary — without
# this, 80% of the executed orders in the 12-08-2026 file would be counted as
# failures.
STATUS_ALIASES = {
    "FILLED": "COMPLETE",
    "COMPLETED": "COMPLETE",
    "EXECUTED": "COMPLETE",
    "CANCELED": "CANCELLED",
}


def _normalise_status(value):
    text = str(value or "").strip().upper()
    return STATUS_ALIASES.get(text, text)


# One column can carry two orderings — 55,659 rows of the 12-08-2026 file read
# "12-08-2026 09:16:00" and 14,101 read "09:15:23 12-08-2026", as if two source
# systems were concatenated. The dedup key uses the timestamp, so both are
# rewritten to "DD-MM-YYYY HH:MM:SS" or the same order would not collapse.
_TS_DATE_FIRST = re.compile(r"^(\d{2}[-/]\d{2}[-/]\d{4})[ T]+(\d{1,2}:\d{2}(?::\d{2})?)")
_TS_TIME_FIRST = re.compile(r"^(\d{1,2}:\d{2}(?::\d{2})?)[ T]+(\d{2}[-/]\d{2}[-/]\d{4})")


def _normalise_timestamp(value):
    text = " ".join(str(value or "").strip().split())
    if not text:
        return ""
    match = _TS_DATE_FIRST.match(text)
    if match:
        return f"{match.group(1).replace('/', '-')} {match.group(2)}"
    match = _TS_TIME_FIRST.match(text)
    if match:
        return f"{match.group(2).replace('/', '-')} {match.group(1)}"
    return text


# Outliers are judged per (date, algo, type) group by ROBUST deviation: a
# user is "in range" when their lots per Cr lies within median +/- k * MAD of
# the group's values (MAD = median absolute deviation, scaled by 1.4826 so it
# matches a std dev on well-behaved data). Unlike mean/std-dev, a blowing
# account cannot widen the range that is meant to catch it. Functions that
# take a `std_multiplier` accept k as a Decimal (e.g. 1, 1.5, 2); default 1.
DEFAULT_STD_MULTIPLIER = Decimal("1")

# The MAD -> std-dev consistency constant, 1 / Phi^-1(0.75) ~ 1.4826.
#
# DERIVED, not typed in: on normal data the median absolute deviation lands at
# 0.6745 sigma, so raw MAD understates spread by a third. Scaling by this
# recovers sigma, which is what makes the `k` in "median +/- k*MAD" mean the
# same thing a std-dev multiplier would. It is a mathematical constant, not a
# tunable — the tunable is k, which is a dashboard input.
#
# Computing it from the normal quantile keeps the reason visible in the code
# rather than in a comment beside a magic number. (The literal 1.4826 changes
# no flag on real data — the relative difference is 1.5e-6 — so this is for
# readability, not accuracy.)
#
# NOTE the assumption it carries: the 0.6745 relationship holds for a NORMAL
# distribution. Where an algo's users are near-identical the spread is tiny and
# "std-dev equivalent" is a loose description, so read such a band as "distance
# from the median" rather than a true sigma.
_MAD_SCALE = Decimal(str(1 / NormalDist().inv_cdf(0.75)))


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------

def _norm_col(value):
    return re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())


def _pick_column(columns, *candidates):
    by_norm = {_norm_col(col): col for col in columns}
    for candidate in candidates:
        found = by_norm.get(_norm_col(candidate))
        if found:
            return found
    return None


def _decimal(value, default="0"):
    try:
        if value is None or str(value).strip().lower() in {"", "none", "nan", "na"}:
            return Decimal(default)
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return Decimal(default)


# Excel number format with Indian digit grouping (16,44,536 / 5,72,58,680);
# the escaped-comma groups drop automatically for smaller magnitudes, and the
# first two conditional sections keep crore-scale values grouped on both signs.
INDIAN_XLSX_FMT = (r"[>=10000000]##\,##\,##\,##\,##0;"
                   r"[<=-10000000]-##\,##\,##\,##\,##0;"
                   r"##\,##\,##0")


def format_indian(value, places=0):
    """Indian-style digit grouping: 1644536 -> "16,44,536" (last three
    digits, then groups of two)."""
    text = f"{float(value):.{places}f}"
    sign = "-" if text.startswith("-") else ""
    text = text.lstrip("-")
    int_part, _, frac = text.partition(".")
    if len(int_part) > 3:
        head, tail = int_part[:-3], int_part[-3:]
        groups = []
        while len(head) > 2:
            groups.insert(0, head[-2:])
            head = head[:-2]
        if head:
            groups.insert(0, head)
        int_part = ",".join(groups + [tail])
    return sign + int_part + (("." + frac) if frac else "")


def format_crore(value):
    """Compact Indian units: 13418000000 -> "1,341.8 Cr", 4038000000 ->
    "403.8 Cr"; below a crore -> lakhs ("53.5 L"); below a lakh -> plain
    Indian-grouped number."""
    v = float(value)
    magnitude = abs(v)
    if magnitude >= 1e7:
        scaled, unit = v / 1e7, " Cr"
    elif magnitude >= 1e5:
        scaled, unit = v / 1e5, " L"
    else:
        return format_indian(v)
    text = format_indian(scaled, places=2)
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text + unit


def _fmt_decimal(value, places=2):
    quant = Decimal("1") if places == 0 else Decimal("1").scaleb(-places)
    try:
        return f"{_decimal(value).quantize(quant):f}"
    except Exception:
        return str(value or "")


def _classify_symbol(symbol):
    upper = str(symbol or "").upper()
    if "BANKNIFTY" in upper or "BANK NIFTY" in upper:
        return "BANKNIFTY"
    if "SENSEX" in upper:
        return "SENSEX"
    if "NIFTY" in upper:
        return "NIFTY"
    return ""


# The SAME option contract circulates under several symbol formats (different
# brokers / servers), so all of them must normalise to one key or a contract
# would be counted as two distinct strikes. Every format seen in the wild:
#
#   spaced, 4-digit year, option type BEFORE strike
#     NIFTY 27JAN2026 PE 25350        -> NIFTY 25350 PE
#   spaced, 2-digit year, strike BEFORE option type
#     SENSEX 16JUL26 77600 CE         -> SENSEX 77600 CE
#   compact, day + month name + 2-digit year
#     NIFTY21JUL2624100PE             -> NIFTY 24100 PE
#   compact monthly, 2-digit year + month name, NO day
#     NIFTY26JAN26400CE               -> NIFTY 26400 CE
#   compact, 2-digit year + month code + day (O/N/D = Oct/Nov/Dec)
#     NIFTY2612024450PE               -> NIFTY 24450 PE
#     SENSEX25O2388700CE              -> SENSEX 88700 CE
#
# The EXPIRY is deliberately NOT parsed from the symbol: one orderbook covers
# a single session in which every symbol of an index shares the same expiry
# (all NIFTY rows one date, all SENSEX rows one date, …), and that date is
# ENTERED in the dashboard — so a symbol only needs to give up its index,
# strike and option type, and the date ambiguities of the compact forms
# (NIFTY26JUL… = July-2026 monthly, not 26-Jul) disappear entirely.
#
# The strike is found by VALUE: index strikes sit in a known band, so in the
# compact forms the tail digits that read as a plausible level are the strike
# (NIFTY26JUL22600CE -> 22600, never the 600 of a "26-Jul-22" misread). The
# bands are deliberately far wider than today's levels (NIFTY ~20-35k,
# SENSEX ~65-88k, BANKNIFTY ~40-60k) to tolerate index drift. Strikes are
# 5 digits today and the 5-wide tail is tried first; should an index ever
# cross 1,00,000 the compact form turns genuinely ambiguous (the 5-digit
# tail of a 6-digit strike can itself land in band) and this rule needs
# revisiting.
# How far from the day's own High / Low a strike may sit and still be accepted.
# Index options list strikes well outside the day's range, so the window is
# deliberately generous — it exists only to reject a MIS-PARSE (reading
# "1124050" out of NIFTY2681124050PE instead of "24050"), not to judge whether
# a strike is sensibly priced. A mis-parse is out by an order of magnitude, so
# nothing this wide can let one through.
STRIKE_BAND_LOW = 0.5
STRIKE_BAND_HIGH = 1.5

_COMPACT_TAIL = re.compile(r"(\d+)(CE|PE)$")


def strike_bands(levels):
    """{index: (low, high)} derived from the FETCHED day High / Low.

    `levels` is marketdata.fetch_index_levels()'s first return value. There is
    no hardcoded rupee range: an index drifts, and a fixed band silently starts
    rejecting real strikes (or accepting mis-parses) as it does. The day's own
    High / Low is the correct reference because every strike traded that day
    sits near it.

    An index with no fetched level gets no band and falls back to the
    self-derived reference in parse_strike."""
    bands = {}
    for name, entry in (levels or {}).items():
        low, high = entry.get("low"), entry.get("high")
        if low and high:
            bands[name] = (float(low) * STRIKE_BAND_LOW,
                           float(high) * STRIKE_BAND_HIGH)
    return bands


def reference_bands_from_symbols(symbols):
    """Fallback when market data is unavailable: derive the band from the
    orderbook itself.

    SPACED symbols are unambiguous — the strike is its own token next to
    CE/PE — so they anchor the scale without any assumption. Their median
    becomes the reference. Only if a file has no spaced symbol at all does
    this return nothing, and parse_strike then takes the widest candidate
    that is a plausible strike step."""
    per_index = {}
    for symbol in symbols:
        text = " ".join(str(symbol or "").upper().split()).replace("BANK NIFTY", "BANKNIFTY")
        segment = _classify_symbol(text)
        if not segment or not text.startswith(segment):
            continue
        tokens = text.split()
        if len(tokens) < 2:
            continue
        for i, tok in enumerate(tokens):
            if tok in ("CE", "PE"):
                for j in (i + 1, i - 1):
                    if 0 <= j < len(tokens) and tokens[j].isdigit():
                        per_index.setdefault(segment, []).append(int(tokens[j]))
    bands = {}
    for segment, values in per_index.items():
        mid = statistics.median(values)
        bands[segment] = (mid * STRIKE_BAND_LOW, mid * STRIKE_BAND_HIGH)
    return bands


def parse_strike(symbol, bands=None):
    """(segment, strike, option type) for an option symbol, else None.

    `bands` ({index: (low, high)}) validates the parse. With no band for the
    index, the compact form falls back to the widest candidate that looks like
    a real strike step — see _plausible_strike."""
    text = " ".join(str(symbol or "").upper().split()).replace("BANK NIFTY", "BANKNIFTY")
    segment = _classify_symbol(text)
    # the index must be the symbol's PREFIX — the substring classifier alone
    # would let FINNIFTY / MIDCPNIFTY options through as NIFTY strikes
    if not segment or not text.startswith(segment):
        return None
    band = (bands or {}).get(segment)

    def ok(value):
        return band is None or band[0] <= value <= band[1]

    tokens = text.split()
    if len(tokens) > 1:
        # spaced forms: the strike is its own token next to CE/PE —
        # "… PE 25350" (after) or "… 77600 CE" (before)
        for i, tok in enumerate(tokens):
            if tok in ("CE", "PE"):
                for j in (i + 1, i - 1):
                    if (0 <= j < len(tokens) and tokens[j].isdigit()
                            and ok(int(tokens[j]))):
                        return (segment, int(tokens[j]), tok)
    # compact forms: the strike is the tail of the digit run before CE/PE.
    # Strikes are 5 digits today; 6 and 4 keep the parse working if an index
    # crosses 1,00,000 or falls below 10,000.
    match = _COMPACT_TAIL.search(text.replace(" ", ""))
    if match:
        run, opt = match.groups()
        candidates = [int(run[-w:]) for w in (5, 6, 4) if len(run) >= w]
        if band is not None:
            for value in candidates:
                if ok(value):
                    return (segment, value, opt)
            return None
        value = _plausible_strike(candidates)
        if value is not None:
            return (segment, value, opt)
    return None


# Index strikes are listed on a round step (50 / 100 / 500), so a correctly
# parsed strike is a multiple of at least 50 while a mis-parse — digits of the
# expiry dragged in — almost never is. Used only when no band is available.
_STRIKE_STEP = 50


def _plausible_strike(candidates):
    for value in candidates:
        if value and value % _STRIKE_STEP == 0:
            return value
    return candidates[0] if candidates else None


def _user_key(value):
    """Canonical user-id key for matching across files. Excel/CSV strips
    leading zeros from numeric ids ("04101961" becomes 4101961), so all-digit
    ids are keyed without leading zeros; matching is also case-insensitive.

    An all-numeric id can also come back from Excel as a FLOAT ("6954037.0"):
    once a sheet holding only numeric ids is opened and saved, the column is
    typed as a number and pandas renders it with a decimal. That trailing
    ".0" is a formatting artifact, never part of an id, so it is dropped
    before the leading-zero rule runs — otherwise "06954037" (text, in one
    file) and "6954037.0" (float, in another) would key differently and the
    account would silently go unmatched. Only a purely zero fraction is
    stripped, so a genuine "1.5" is left alone."""
    key = str(value or "").strip().upper()
    if "." in key:
        head, _, tail = key.partition(".")
        if head.isdigit() and tail and set(tail) == {"0"}:
            key = head
    if key.isdigit():
        return key.lstrip("0") or "0"
    return key


def _server_key(value):
    """Canonical server key for matching the orderbook against the MTM
    summary; a missing/NaN server reads as blank."""
    key = str(value or "").strip().upper()
    return "" if key in {"NAN", "NONE", "NA"} else key


_DATE_FORMATS = ("%d-%m-%Y", "%Y-%m-%d", "%d-%b-%Y", "%d/%m/%Y")
_DATE_PARSE_CACHE = {}


def _parse_date(value):
    raw = str(value or "").strip().split(" ")[0]
    if not raw:
        return None
    if raw in _DATE_PARSE_CACHE:
        return _DATE_PARSE_CACHE[raw]
    parsed = None
    for fmt in _DATE_FORMATS:
        try:
            parsed = datetime.strptime(raw, fmt).date()
            break
        except ValueError:
            continue
    _DATE_PARSE_CACHE[raw] = parsed
    return parsed


# ---------------------------------------------------------------------------
# Lot-size history (divisor applied to the traded quantity)
# ---------------------------------------------------------------------------

def _nifty_lot_size(d):
    # up to 26-Dec-2024 -> 25; 27-Dec-2024..30-Dec-2025 -> 75
    # (29/30-Jan-2025 -> 25, exception inside the 75 window); from 31-Dec-2025 -> 65
    if d is None:
        return Decimal("65")
    if d <= date(2024, 12, 26) or d in (date(2025, 1, 29), date(2025, 1, 30)):
        return Decimal("25")
    if d <= date(2025, 12, 30):
        return Decimal("75")
    return Decimal("65")


def _sensex_lot_size(d):
    # up to 03-Jan-2025 -> 10 (also 30/31-Jan-2025); otherwise -> 20
    if d is None:
        return Decimal("20")
    if d <= date(2025, 1, 3) or d in (date(2025, 1, 30), date(2025, 1, 31)):
        return Decimal("10")
    return Decimal("20")


def _banknifty_lot_size(d):
    # up to 26-Feb-2025 -> 15; 27-Feb-2025..26-Jun-2025 -> 30;
    # 27-Jun-2025..30-Dec-2025 -> 35; from 31-Dec-2025 -> 30
    if d is None:
        return Decimal("30")
    if d <= date(2025, 2, 26):
        return Decimal("15")
    if d <= date(2025, 6, 26):
        return Decimal("30")
    if d <= date(2025, 12, 30):
        return Decimal("35")
    return Decimal("30")


_LOT_SIZE_BY_SEGMENT = {
    "NIFTY": _nifty_lot_size,
    "SENSEX": _sensex_lot_size,
    "BANKNIFTY": _banknifty_lot_size,
}


# ---------------------------------------------------------------------------
# Generic table reading: CSV or Excel, from a path or an uploaded file object
# ---------------------------------------------------------------------------

def _cell_str(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d-%m-%Y %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%d-%m-%Y")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _excel_rows(source):
    from openpyxl import load_workbook
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        # First sheet holds the per-row data (later sheets are summary views).
        for row in workbook.worksheets[0].iter_rows(values_only=True):
            yield [_cell_str(v) for v in row]
    finally:
        workbook.close()


def _csv_rows(source):
    if hasattr(source, "read"):
        raw = source.read()
        text = raw.decode("utf-8-sig", errors="replace") if isinstance(raw, bytes) else raw
        fh = io.StringIO(text)
    else:
        fh = open(source, newline="", encoding="utf-8-sig", errors="replace")
    try:
        yield from csv.reader(fh)
    finally:
        fh.close()


def _read_table(source, name=None):
    """Return (headers, row_iterator) for a CSV or Excel table. `source` may be
    a filesystem path or a file-like object (e.g. a Streamlit upload); `name`
    supplies the filename when `source` has none."""
    filename = str(name or getattr(source, "name", source)).lower()
    rows = _excel_rows(source) if filename.endswith((".xlsx", ".xlsm", ".xls")) else _csv_rows(source)
    try:
        headers = next(rows)
    except StopIteration:
        return [], iter(())
    return headers, rows


def _cell(row, idx):
    if idx is None or idx >= len(row):
        return ""
    return (row[idx] or "").strip()


# ---------------------------------------------------------------------------
# Reading the two inputs
# ---------------------------------------------------------------------------

Order = namedtuple(
    "Order",
    "rowid trade_date server user_id segment qty avg_price order_id exch_order_id symbol "
    "category status minute order_time exchange_time tag",
)

# "09:16" from "05-Aug-2026 09:16:00" / "09:16:00" / a datetime cell. The
# intraday chart buckets on this, so only hours and minutes are kept.
_TIME_RE = re.compile(r"(\d{1,2}):(\d{2})")


def _parse_minute(value):
    match = _TIME_RE.search(str(value or ""))
    if not match:
        return ""
    hour, minute = int(match.group(1)), int(match.group(2))
    if 0 <= hour <= 23 and 0 <= minute <= 59:
        return f"{hour:02d}:{minute:02d}"
    return ""

# Order tags mark the trade kind: "h_..." = hedge, "v_..." = VAR, "s_..." =
# square-off; anything else is a normal trade.
ORDER_CATEGORIES = ("normal", "hedge", "var", "sqoff")


def _order_category(tag):
    text = str(tag or "").strip().lower()
    if text.startswith("h_"):
        return "hedge"
    if text.startswith("v_"):
        return "var"
    if text.startswith("s_"):
        return "sqoff"
    return "normal"


FO_EXCHANGES = {"BFO", "NFO"}


def read_orderbook(source, name=None, all_statuses=False):
    """Parse an orderbook CSV/Excel into Order tuples. Rows that never traded
    (status != COMPLETE), sit on a non-F&O exchange (anything but BFO / NFO —
    e.g. NSE cash rows like NIFTYBEES-EQ) or belong to none of the three
    segments are dropped.

    `all_statuses=True` keeps the non-COMPLETE rows (rejected / cancelled /
    failed) as well — they carry no traded quantity, so they feed only the
    orders summary, never the trade value or strike math."""
    headers, rows = _read_table(source, name)
    if not headers:
        return []

    def col(*candidates):
        found = _pick_column(headers, *candidates)
        return headers.index(found) if found else None

    # Both layouts are auto-detected: the compiled orderbook ("User ID",
    # "Symbol", "Tag") and the newer export ("user_id", "traiding_symbol",
    # "order_unique_identifier"). Candidate order matters — the first match
    # wins — so where a file carries BOTH spellings the wanted one is listed
    # first.
    i_date = col("date", "_date", "trade_date")
    i_server = col("server")
    i_user = col("user_id", "userid", "user id", "UserID")
    i_exchange = col("exchange", "exch", "exchg")
    # note the misspelling in the newer export: "traiding_symbol"
    i_symbol = col("symbol", "traiding_symbol", "trading_symbol", "tradingsymbol")
    i_avg = col("avg_price", "avg price", "avg_traded_price",
                "OrderAverageTradedPrice", "AveragePrice", "price")
    # order_quantity is the quantity PLACED; `quantity` in the newer export is
    # the quantity FILLED, which collapses a cancelled order to near zero.
    # "lots fired" means placed, so order_quantity is preferred where both exist.
    i_qty = col("order_quantity", "quantity", "qty", "OrderQuantity")
    i_status = col("status", "order_status", "OrderStatus")
    i_order_id = col("order_id", "orderid", "order id", "broker_order_id")
    i_exch_order_id = col("exchg_order_id", "exch_order_id", "exchange_order_id",
                          "exchgorderid")
    i_rowid = col("row_id", "sno", "id")
    i_tag = col("tag", "order_unique_identifier")
    # the clock time the order was placed — feeds the intraday chart only, so a
    # file without it simply produces no dots rather than failing
    i_time = col("order time", "order_time", "order_generated_time", "time")
    # the dedup key needs the exchange timestamp too; a file without it simply
    # contributes a blank component rather than failing
    i_exch_time = col("exchange time", "exchange_time", "exchange_transact_time",
                      "exchg time")
    if i_symbol is None or i_avg is None or i_qty is None:
        raise ValueError("orderbook is missing the symbol / avg price / quantity columns")

    if i_time is None or i_exch_time is None:
        logger.warning("orderbook has no %s column — the dedup key loses that "
                       "component and may merge distinct orders",
                       "Order Time" if i_time is None else "Exchange Time")

    orders = []
    for line_no, raw in enumerate(rows, start=2):
        if i_exchange is not None and _cell(raw, i_exchange).upper() not in FO_EXCHANGES:
            continue
        status = (_normalise_status(_cell(raw, i_status))
                  if i_status is not None else "COMPLETE")
        if not all_statuses and status != "COMPLETE":
            continue
        symbol = _cell(raw, i_symbol)
        segment = _classify_symbol(symbol)
        if not segment:
            continue
        try:
            rowid = int(_cell(raw, i_rowid))
        except (ValueError, TypeError):
            rowid = line_no
        raw_tag = _cell(raw, i_tag)
        raw_order_time = _normalise_timestamp(_cell(raw, i_time))
        orders.append(Order(
            rowid=rowid,
            trade_date=_parse_date(_cell(raw, i_date)),
            server=sys.intern(_cell(raw, i_server)),
            user_id=sys.intern(_cell(raw, i_user)),
            segment=segment,
            qty=_decimal(_cell(raw, i_qty)),
            avg_price=_decimal(_cell(raw, i_avg)),
            order_id=_cell(raw, i_order_id),
            exch_order_id=_cell(raw, i_exch_order_id),
            symbol=sys.intern(symbol),
            category=_order_category(raw_tag),
            status=sys.intern(status),
            minute=sys.intern(_parse_minute(raw_order_time)),
            order_time=sys.intern(raw_order_time),
            exchange_time=sys.intern(_normalise_timestamp(_cell(raw, i_exch_time))),
            tag=sys.intern(raw_tag),
        ))
    return orders


def fill_missing_servers(orders, allocations, label="orderbook"):
    """Backfill a blank SERVER from the User MTM, for exports that omit it.

    The newer orderbook has no server column at all, but the MTM does and the
    order's user id matches it, so the server can be recovered per user.

    Only filled when the user has EXACTLY ONE server in the MTM. A user id
    that exists on two servers is two distinct accounts with their own algo and
    allocation, and guessing between them is precisely what `_pick_allocation`
    refuses to do — those orders keep their blank server and stay unmatched,
    which is visible rather than wrong. `aliases.json` is keyed on
    (id, server) and therefore cannot resolve them either; the count is logged.

    Returns (orders, report). Orders already carrying a server are untouched.
    """
    by_user = {}
    for key, entries in (allocations or {}).items():
        servers = {e["server"] for e in entries if e["server"]}
        if len(servers) == 1:
            by_user[key] = servers.pop()

    filled = ambiguous = already = 0
    out = []
    for row in orders:
        if _server_key(row.server):
            already += 1
            out.append(row)
            continue
        server = by_user.get(_user_key(row.user_id))
        if server:
            filled += 1
            out.append(row._replace(server=sys.intern(server)))
        else:
            ambiguous += 1
            out.append(row)

    report = {"already": already, "filled": filled, "unresolved": ambiguous}
    if filled or ambiguous:
        logger.info("%s: server backfilled from the User MTM for %d order(s); "
                    "%d already had one; %d could not be resolved (user absent "
                    "from the MTM, or present on more than one server)",
                    label, filled, already, ambiguous)
    return out, report


def indexes_in_orderbook(orders):
    """The indexes the orderbook actually traded, in report order.

    Used to fetch market data for those alone — BANKNIFTY is absent on most
    days, and fetching it anyway costs a network round trip and produces a
    spurious "no data" warning."""
    seen = {row.segment for row in orders if row.segment}
    return [name for name in ("NIFTY", "BANKNIFTY", "SENSEX") if name in seen]


def read_allocations(source, name=None):
    """Read the compiled user MTM summary and return a lookup keyed by the
    canonical user id (see _user_key): {key: [{"allocation", "algo",
    "user_id", "server"}, ...]}. The same user id can legitimately run on
    more than one server — a distinct account with its own algo and
    allocation — so one entry is kept per (user, server) row and aggregate()
    resolves duplicates by the orderbook row's server. The original user id
    string is kept so the report can show the proper zero-padded form when
    the orderbook lost the leading zeros."""
    headers, rows = _read_table(source, name)
    if not headers:
        return {}
    user_col = _pick_column(headers, "user_id", "userid", "user id", "UserID")
    alloc_col = _pick_column(headers, "allocation")
    algo_col = _pick_column(headers, "algo")
    server_col = _pick_column(headers, "server")
    if user_col is None or alloc_col is None:
        raise ValueError("summary is missing the UserID / ALLOCATION columns")
    i_user = headers.index(user_col)
    i_alloc = headers.index(alloc_col)
    i_algo = headers.index(algo_col) if algo_col else None
    i_server = headers.index(server_col) if server_col else None

    allocations = {}
    for raw in rows:
        user = _cell(raw, i_user)
        key = _user_key(user)
        if not key:
            continue
        server = _server_key(_cell(raw, i_server)) if i_server is not None else ""
        entries = allocations.setdefault(key, [])
        if any(e["server"] == server for e in entries):
            continue  # true duplicate (same user AND server) — keep the first
        entries.append({
            "allocation": _decimal(_cell(raw, i_alloc)),
            "algo": _cell(raw, i_algo),
            "user_id": user,
            "server": server,
        })
    return allocations


def _type_index(type_map):
    """{user key: its one Int / Pos+Int classification} for users whose every
    account agrees — the fallback tier for _resolve_type."""
    seen = {}
    for (user_key, _server), user_type in (type_map or {}).items():
        seen.setdefault(user_key, set()).add(user_type)
    return {user: next(iter(types)) for user, types in seen.items() if len(types) == 1}


def _resolve_type(type_map, type_index, user_key, server_key):
    """An account's Int / Pos+Int, resolved the way _pick_allocation resolves
    its allocation: the exact (user, server) account first, then the user's
    own classification when all of their accounts agree.

    Crucially it NEVER invents a default. A plain type_map.get(..., "Int")
    fabricates an Intraday account whenever the server labels disagree
    between the orderbook and the MTM — which is how an algo whose accounts
    are every one of them Positional grew a phantom Int row. An account the
    segregation genuinely doesn't know stays blank instead."""
    exact = (type_map or {}).get((user_key, server_key))
    if exact is not None:
        return exact
    return type_index.get(user_key, "")


ACCOUNT_ALIAS_FILE = SCRIPT_DIR / "aliases.json"


def load_account_aliases(path=None):
    """{(orderbook user key, server key): MTM account id} from
    aliases.json.

    Keyed on the SERVER as well as the id, because one base id is a different
    account per server — TB2433 is TB2433A41 on VS8 and TB2433A42 on VS29.
    Missing file -> empty map, and matching falls back to prefix inference."""
    import json

    path = Path(path) if path else ACCOUNT_ALIAS_FILE
    if not path.exists():
        return {}
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError) as exc:
        logger.error("aliases.json could not be read (%s) — falling back "
                     "to prefix inference", exc)
        return {}
    out = {}
    for base, per_server in raw.items():
        if str(base).startswith("_") or not isinstance(per_server, dict):
            continue                       # the "_comment" block
        for server, account in per_server.items():
            out[(_user_key(base), _server_key(server))] = str(account)
    return out


def load_id_aliases(path=None):
    """{orderbook user key: MTM account id} from aliases.json's STRING entries.

    The object entries are per-server (one base id is a different account on
    each server). These are not: they say "this id IS that MTM account,
    wherever it traded". Two things need that:

      * typos in the orderbook — S03939TWO is S0393TWO with a stray 9, and it
        fired on VS5 while the real account lives on VS2, so no per-server
        entry could express it;
      * exports that write the All User id (CC03) where the MTM uses its own
        (XLDH161), which otherwise drops every one of that account's orders.

    Deliberately server-blind, so it is only safe for ids stated by hand.
    Nothing is inferred by similarity: a fuzzy rule would eventually fold two
    real accounts together and move their orders under the wrong algo."""
    import json

    path = Path(path) if path else ACCOUNT_ALIAS_FILE
    if not path.exists():
        return {}
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError) as exc:
        logger.error("aliases.json could not be read (%s)", exc)
        return {}
    return {_user_key(k): str(v) for k, v in raw.items()
            if not str(k).startswith("_") and isinstance(v, str)}


def _apply_id_aliases(allocations, orders, id_aliases):
    """Attach stated id -> MTM account mappings, whatever the server."""
    if not id_aliases:
        return {}
    by_key = {}
    for entries in allocations.values():
        for entry in entries:
            by_key.setdefault(_user_key(entry["user_id"]), entry)

    seen = {_user_key(o.user_id) for o in orders}
    resolved = {}
    for key, account in id_aliases.items():
        if key not in seen or key in allocations:
            continue                       # absent today, or matched on its own id
        entry = by_key.get(_user_key(account))
        if entry is None:
            logger.warning("aliases: %s -> %s, but %s is not in the User MTM "
                           "— skipped", key, account, account)
            continue
        allocations[key] = [entry]
        resolved[key] = entry["user_id"]
        logger.info("aliases: %s -> %s (%s, algo %s)", key, entry["user_id"],
                    entry["server"], entry["algo"] or "-")
    return resolved


def add_user_aliases(allocations, orders, account_aliases=None):
    """See below. `account_aliases` (from load_account_aliases) is applied
    FIRST; anything it does not cover falls back to prefix inference."""
    explicit = (load_account_aliases() if account_aliases is None
                else account_aliases)
    resolved = _apply_explicit_aliases(allocations, orders, explicit)
    resolved.update(_apply_id_aliases(allocations, orders, load_id_aliases()))
    resolved.update(_infer_user_aliases(allocations, orders))
    return resolved


def _apply_explicit_aliases(allocations, orders, explicit):
    """Attach the stated (id, server) -> account mappings. An entry whose
    account is not in the User MTM is logged and skipped rather than inventing
    an allocation."""
    if not explicit:
        return {}
    by_account = {}
    for entries in allocations.values():
        for entry in entries:
            by_account[(_user_key(entry["user_id"]), entry["server"])] = entry

    seen = {(_user_key(o.user_id), _server_key(o.server)) for o in orders}
    resolved = {}
    for (key, server), account in explicit.items():
        if (key, server) not in seen:
            continue                       # not trading today — nothing to do
        entry = by_account.get((_user_key(account), server))
        if entry is None:
            logger.warning("account_aliases: %s on %s -> %s, but %s is not in "
                           "the User MTM for that server — skipped",
                           key, server, account, account)
            continue
        bucket = allocations.setdefault(key, [])
        if any(e["server"] == server for e in bucket):
            continue                       # already matched on its own id
        bucket.append(entry)
        resolved[(key, server)] = entry["user_id"]
        logger.info("account_aliases: %s on %s -> %s (algo %s)",
                    key, server, entry["user_id"], entry["algo"] or "-")
    return resolved


def _infer_user_aliases(allocations, orders):
    """Match orderbook ids that drop the MTM's account suffix, in place.

    The orderbook writes the BASE user id while the MTM writes the account:
    orderbook JSR129 on VS5 is the MTM's JSR129A31, and on VS26 it is
    JSR129A25 — two distinct accounts under one base id, told apart by the
    server. Left unmatched these users get no algo and no type at all, so
    they vanish from the Algo Summary and pile up in the "—" row while their
    lots still count in the KPI.

    Stripping a trailing "A<digits>" is NOT safe — R7RA1315 would reduce to
    R7R and 7RA110119 to 7R, colliding with real ids. So the match is by
    PREFIX WITHIN THE SAME SERVER, and only when exactly one MTM account on
    that server extends the orderbook id. Anything ambiguous is left
    unmatched rather than guessed.

    Returns {(user key, server): matched MTM user id} for reporting."""
    by_server = {}
    for entries in allocations.values():
        for entry in entries:
            by_server.setdefault(entry["server"], []).append(entry)

    unmatched = set()
    for row in orders:
        key = _user_key(row.user_id)
        if _pick_allocation(allocations, key, row.server) is None:
            unmatched.add((key, _server_key(row.server)))

    resolved = {}
    for key, server in sorted(unmatched):
        if not key:
            continue
        candidates = [e for e in by_server.get(server, ())
                      if len(e["user_id"]) > len(key)
                      and _user_key(e["user_id"]).startswith(key)]
        if len(candidates) != 1:
            continue  # none, or ambiguous — never guess
        entry = candidates[0]
        bucket = allocations.setdefault(key, [])
        if any(e["server"] == server for e in bucket):
            continue
        bucket.append(entry)
        resolved[(key, server)] = entry["user_id"]
    return resolved


def _pick_allocation(allocations, user_key, server):
    """Resolve a user's MTM entry: unambiguous when the user has one row;
    a user id that appears on several servers (distinct accounts) is matched
    on the orderbook row's server, and stays unmatched (blank allocation)
    when no server agrees rather than guessing the wrong account."""
    entries = allocations.get(user_key)
    if not entries:
        return None
    if len(entries) == 1:
        return entries[0]
    server = _server_key(server)
    return next((e for e in entries if e["server"] == server), None)


# ---------------------------------------------------------------------------
# Dedup, aggregation and output
# ---------------------------------------------------------------------------

# Order ids that went through Excel can come back in scientific notation
# ("2.60623E+13"), which collapses many DISTINCT orders onto one value.
# Such ids cannot be used as a dedup key.
_SCI_NOTATION = re.compile(r"^-?\d+(\.\d+)?E[+-]?\d+$", re.IGNORECASE)

# How many colliding keys to keep for the on-screen report.
DUP_SAMPLE_SIZE = 5


def duplicate_report(keys, label, sample_size=DUP_SAMPLE_SIZE):
    """Count rows sharing a key, log the result, and return it as a dict.

    `keys` is the per-row key sequence. Returns {"label", "rows", "distinct",
    "colliding", "surplus", "samples"} where `surplus` is how many rows a
    dedup on this key would remove.

    Logged at INFO when the file is clean and WARNING when it is not, so a
    duplicated upload is visible in the log without reading the dashboard."""
    counts = Counter(keys)
    colliding = [k for k, n in counts.items() if n > 1]
    surplus = sum(counts[k] - 1 for k in colliding)
    report = {
        "label": label,
        "rows": len(keys),
        "distinct": len(counts),
        "colliding": len(colliding),
        "surplus": surplus,
        "samples": [(k, counts[k]) for k in colliding[:sample_size]],
    }
    if colliding:
        logger.warning(
            "%s: %d duplicate row(s) across %d key(s) — %d of %d rows are "
            "surplus and will be dropped (sample: %s)",
            label, surplus, len(colliding), surplus, len(keys),
            ", ".join(f"{k}x{n}" for k, n in report["samples"]),
        )
    else:
        logger.info("%s: no duplicates — %d rows, all keys distinct",
                    label, len(keys))
    return report


def _order_key(row):
    """The orderbook's uniqueness key:

        User ID + Order ID + Order Time + Exchg Order ID + Exchange Time + Tag

    Six components rather than four because a compiled orderbook that has been
    through Excel cannot be keyed on Order ID alone. In the 11-08-2026 file
    **21% of Order IDs came back in scientific notation** ("2.60811E+13"),
    which collapses thousands of distinct orders onto one value. The earlier
    key (user + date + order id + symbol) had to skip those rows entirely,
    so 16,519 genuine duplicates among them were counted twice.

    The exchange fields survive Excel intact, so including Exchg Order ID and
    both timestamps keeps the key discriminating even when Order ID is
    corrupted. Measured on 11-08: this key collapses 228,722 rows, of which
    228,719 are byte-identical duplicates — it removes duplicates and almost
    nothing else.
    """
    return (_user_key(row.user_id), row.order_id, row.order_time,
            row.exch_order_id, row.exchange_time, row.tag)


def dedup_orders(rows, label="orderbook"):
    """Drop duplicate orders on the six-part key above, keeping the first
    occurrence (lowest row id). The duplicate count is logged; use
    dedup_orders_with_report when the caller also wants to display it."""
    return dedup_orders_with_report(rows, label)[0]


def dedup_orders_with_report(rows, label="orderbook"):
    """(deduped rows, duplicate report) — see duplicate_report.

    Every row can be keyed: the key no longer depends on Order ID being
    intact, so there is no bypass path. A row whose Order ID is corrupted is
    still separated by its exchange id and timestamps."""
    best = {}
    keys = []
    corrupt = 0
    for row in rows:
        if row.order_id and _SCI_NOTATION.match(row.order_id):
            corrupt += 1
        key = _order_key(row)
        keys.append(key)
        current = best.get(key)
        if current is None or row.rowid < current.rowid:
            best[key] = row
    report = duplicate_report(keys, label)
    report["unkeyable"] = 0
    report["corrupt_order_ids"] = corrupt
    if corrupt:
        logger.warning(
            "%s: %d row(s) carry an Order ID mangled into scientific notation "
            "by Excel — the key still separates them by Exchg Order ID and the "
            "timestamps, but the source file should be exported as text",
            label, corrupt)
    return list(best.values()), report


def aggregate(rows, allocations=None, std_multiplier=None, type_map=None):
    """Group orders by (date, server, user, segment) and attach each user's
    allocation from the summary (None when the user isn't in the summary).
    Users are grouped on their canonical id, so "4101961" and "04101961"
    are the same user; the report shows the zero-padded form when known.

    `type_map` classifies each (user key, server key) account as "Int" or
    "Pos+Int" (the segregation classification); the Lots per Cr statistics
    and outlier bands are then computed per (date, algo, type) group.
    Without a map every account is one unlabelled group per algo."""
    allocations = allocations or {}
    type_map = type_map or {}
    type_index = _type_index(type_map)
    groups = {}
    display = {}
    for row in rows:
        user_key = _user_key(row.user_id)
        shown = display.get(user_key)
        if shown is None or (row.user_id.isdigit() and len(row.user_id) > len(shown)):
            display[user_key] = row.user_id
        key = (row.trade_date, row.server, user_key, row.segment)
        group = groups.get(key)
        if group is None:
            group = groups[key] = {
                "order_count": 0,
                "quantity": Decimal("0"),
                "lots": Decimal("0"),
                "trade_value": Decimal("0"),
            }
        lot_size = _LOT_SIZE_BY_SEGMENT[row.segment](row.trade_date)
        group["order_count"] += 1
        group["quantity"] += row.qty
        group["lots"] += abs(row.qty) / lot_size
        group["trade_value"] += row.avg_price * row.qty

    out = []
    for (trade_date, server, user_key, segment), group in groups.items():
        entry = _pick_allocation(allocations, user_key, server)
        user_id = display[user_key]
        if entry:
            # Prefer the summary's zero-padded id over a leading-zero-stripped one.
            mtm_id = entry["user_id"]
            if mtm_id.isdigit() and user_id.isdigit() and len(mtm_id) > len(user_id):
                user_id = mtm_id
        out.append({
            "trade_date": trade_date,
            "server": server,
            "user_id": user_id,
            "allocation": entry["allocation"] if entry else None,
            "algo": entry["algo"] if entry else "",
            # type comes from the MATCHED MTM account, whose id can carry an
            # account suffix the orderbook drops (JSR129 -> JSR129A31)
            "user_type": _resolve_type(
                type_map, type_index,
                _user_key(entry["user_id"]) if entry else user_key,
                _server_key(server)),
            "segment": segment,
            "order_count": group["order_count"],
            "quantity": group["quantity"],
            "lots": group["lots"],
            "trade_value": group["trade_value"],
        })
    # date DESC, then server, user_id, segment ASC
    out.sort(key=lambda r: (r["server"], r["user_id"], r["segment"]))
    out.sort(key=lambda r: r["trade_date"] or date.min, reverse=True)
    return _attach_lot_metrics(out, std_multiplier)


def _lots_per_cr_stats(rows, std_multiplier=None):
    """Per (date, algo, type) ROBUST statistics of lots per Cr: the median,
    the scaled MAD (1.4826 × median absolute deviation — a std-dev
    equivalent that a blowing account cannot inflate), and the outlier range
    median +/- k * MAD. Returns {(date, algo, type): (median, mad, low, high)}."""
    k = std_multiplier if std_multiplier is not None else DEFAULT_STD_MULTIPLIER
    groups = {}
    for row in rows:
        if row.get("lots_per_cr") is None or not row["algo"]:
            continue
        groups.setdefault((row["trade_date"], row["algo"], row.get("user_type", "")),
                          []).append(row["lots_per_cr"])

    stats = {}
    for key, values in groups.items():
        med = statistics.median(values)
        mad = statistics.median(abs(v - med) for v in values) * _MAD_SCALE
        stats[key] = (med, mad, med - k * mad, med + k * mad)
    return stats


def user_lot_observations(rows, std_multiplier=None):
    """One observation per (date, algo, user): the OUTLIER UNIT IS THE USER.

    A report row is per (user, segment), so a user trading NIFTY and SENSEX
    would otherwise be judged twice on partial exposures. Here the user's
    lots are summed across all their segments (and servers) inside the algo
    and normalised by their TOTAL allocation (one account per server):

        normalise   = total allocation / 1,00,000
        lots per Cr = total lots / normalise

    Every account normalises the same way — allocations below 1,00,000 get a
    fractional normalise (80,000 -> 0.8, …, 20,000 -> 0.2). Users with no
    usable allocation carry no metric and a blank flag. Each observation
    carries the (date, algo) band (mean ± k·σ over the per-user lots per Cr)
    and its outlier flag. Rows without an algo produce no observation."""
    obs_map = {}
    for row in rows:
        if not row["algo"]:
            continue
        key = (row["trade_date"], row["algo"], row.get("user_type", ""),
               _user_key(row["user_id"]))
        o = obs_map.get(key)
        if o is None:
            o = obs_map[key] = {
                "trade_date": row["trade_date"],
                "algo": row["algo"],
                "user_type": key[2],
                "user_key": key[3],
                "user_id": row["user_id"],
                "lots": Decimal("0"),
                "_alloc_by_server": {},
                "_servers": [],
            }
        o["lots"] += row["lots"]
        server = _server_key(row["server"])
        if server not in o["_servers"]:
            o["_servers"].append(server)
        if row["allocation"] is not None:
            o["_alloc_by_server"][server] = row["allocation"]

    out = list(obs_map.values())
    for o in out:
        alloc_by_server = o.pop("_alloc_by_server")
        allocation = sum(alloc_by_server.values(), Decimal("0")) if alloc_by_server else None
        o["allocation"] = allocation
        o["normalise"] = (allocation / NORMALISE_BASE) if allocation else None
        o["lots_per_cr"] = (o["lots"] / o["normalise"]) if o["normalise"] else None
        o["server"] = " / ".join(s for s in o.pop("_servers") if s)

    stats = _lots_per_cr_stats(out, std_multiplier)
    for o in out:
        stat = stats.get((o["trade_date"], o["algo"], o["user_type"]))
        if o["lots_per_cr"] is None or stat is None:
            o["outlier"], o["band_low"], o["band_high"] = "", None, None
            continue
        _, _, low, high = stat
        o["band_low"], o["band_high"] = low, high
        if o["lots_per_cr"] < low:
            o["outlier"] = "Below average range"
        elif o["lots_per_cr"] > high:
            o["outlier"] = "Above average range"
        else:
            o["outlier"] = "In range"
    return out


def _attach_lot_metrics(rows, std_multiplier=None):
    """Add Normalise (allocation / 1,00,000) and Lots per Cr (that ROW's lots
    ÷ normalise — it documents the row), then judge outliers PER USER
    (combined lots ÷ combined normalise, see user_lot_observations) and
    stamp the user's flag on each of their rows. Rows with no allocation or
    no algo can't be judged and stay blank."""
    for row in rows:
        allocation = row["allocation"]
        row["normalise"] = (allocation / NORMALISE_BASE) if allocation else None
        row["lots_per_cr"] = (row["lots"] / row["normalise"]) if row["normalise"] else None

    flags = {(o["trade_date"], o["algo"], o["user_type"], o["user_key"]): o["outlier"]
             for o in user_lot_observations(rows, std_multiplier)}
    for row in rows:
        row["outlier"] = flags.get(
            (row["trade_date"], row["algo"], row.get("user_type", ""),
             _user_key(row["user_id"])), "")
    return rows


def split_rows_by_segment(rows, std_multiplier=None):
    """{index: rows} — the report rows grouped per index (NIFTY / BANKNIFTY /
    SENSEX, whichever the orderbook holds), each group's outlier bands
    RE-JUDGED within that index. The split is in place: `rows` keeps the
    corrected flags.

    Lot sizes differ per index (NIFTY 65, SENSEX 20, BANKNIFTY 30 today), so
    lots per Cr sits on a different scale in each — pooling them would give
    every index a median that belongs to none of them.

    A report row is per (user, segment), so each row lands in exactly one
    group. NOTE the trade-off: a user trading two indexes now appears in
    both tables, and in each their lots per Cr is that index's lots over
    their WHOLE allocation — a partial-exposure figure, not the combined
    one. multi_index_users() reports how many users that affects."""
    groups = {}
    for row in rows:
        groups.setdefault(row["segment"], []).append(row)
    for part in groups.values():
        _attach_lot_metrics(part, std_multiplier)
    return dict(sorted(groups.items()))


def multi_index_users(rows):
    """How many distinct users traded more than one index, and which — the
    accounts whose per-index Lots per Cr is a partial exposure (see
    split_rows_by_segment). Returns (count, {user id: [indexes]})."""
    by_user = {}
    for row in rows:
        by_user.setdefault(row["user_id"], set()).add(row["segment"])
    multi = {user: sorted(segs) for user, segs in by_user.items() if len(segs) > 1}
    return len(multi), multi


def format_row(row):
    return [
        row["trade_date"].strftime("%d-%m-%Y") if row["trade_date"] else "",
        row["server"],
        row["user_id"],
        _fmt_decimal(row["allocation"], places=0) if row["allocation"] is not None else "",
        row.get("algo", ""),
        row.get("user_type", ""),
        row["segment"],
        row["order_count"],
        _fmt_decimal(row["quantity"], places=0),
        _fmt_decimal(row["lots"], places=0),
        # every magnitude in a table is a whole number; only the Cr / Lakh
        # compact values (format_crore) and the percentages keep decimals
        _fmt_decimal(row["normalise"], places=0) if row.get("normalise") is not None else "",
        _fmt_decimal(row["lots_per_cr"], places=0) if row.get("lots_per_cr") is not None else "",
        _fmt_decimal(row["trade_value"], places=0),
        row.get("outlier", ""),
    ]


def algo_summary(rows, std_multiplier=None):
    """Per (date, algo) outlier summary, ALL columns in users: "Total Users"
    (distinct users of the algo), the per-USER lots per Cr statistics (median,
    median +/- k*MAD range — see user_lot_observations), and how many USERS
    fall below / in / above the band, so Below + In + Above = Total Users
    (users with no usable allocation carry no lots per Cr and are the only
    ones that can fall outside the three flag columns). Pass the SAME
    std_multiplier that was given to aggregate() so the band shown matches
    the flags on the rows."""
    observations = user_lot_observations(rows, std_multiplier)
    stats = _lots_per_cr_stats(observations, std_multiplier)
    counts = {}
    for o in observations:
        key = (o["trade_date"], o["algo"], o["user_type"])
        group = counts.setdefault(key, {"users": 0, "below": 0, "in_range": 0, "above": 0})
        group["users"] += 1
        if o["outlier"] == "Below average range":
            group["below"] += 1
        elif o["outlier"] == "Above average range":
            group["above"] += 1
        elif o["outlier"] == "In range":
            group["in_range"] += 1

    out = []
    for (trade_date, algo, user_type), group in counts.items():
        # a group whose users all lack an allocation has no statistics
        mean, std, low, high = stats.get((trade_date, algo, user_type),
                                         (None, None, None, None))
        out.append({
            "trade_date": trade_date,
            "algo": algo,
            "user_type": user_type,
            "users": group["users"],
            "median_lots_per_cr": mean,
            "std_dev": std,
            "band_low": low,
            "band_high": high,
            "below": group["below"],
            "in_range": group["in_range"],
            "above": group["above"],
        })
    out.sort(key=lambda r: (
        (0, int(r["algo"])) if str(r["algo"]).isdigit() else (1, str(r["algo"])),
        r["user_type"],
    ))
    out.sort(key=lambda r: r["trade_date"] or date.min, reverse=True)
    return out


def format_summary_row(row):
    has_stats = row["median_lots_per_cr"] is not None
    return [
        row["algo"],
        row["user_type"],
        row["users"],
        _fmt_decimal(row["median_lots_per_cr"], places=0) if has_stats else "",
        (f"{_fmt_decimal(row['band_low'], places=0)}–"
         f"{_fmt_decimal(row['band_high'], places=0)}") if has_stats else "",
        row["below"],
        row["in_range"],
        row["above"],
    ]


def order_summary(orders, allocations=None, type_map=None, bands=None):
    """Per (algo, Int / Pos+Int) order counts, each row carrying its server
    breakdown for the drill-down:

        {"algo", "user_type", "users", "orders", "executed", "failed",
         "hedge", "var", "servers": [{"server", "users", "orders", ...}]}

    `orders` must come from read_orderbook(..., all_statuses=True) — with
    only the COMPLETE rows the Failed column is structurally 0.

      * Total Orders — every order the user fired, whatever its outcome
      * Executed     — status COMPLETE (ties to the Trade Value Orders KPI)
      * Failed       — rejected / cancelled
      * Pending      — still live at end of day (OPEN / OPEN_PENDING): NOT
                       a failure, so it gets its own column
      * Hedge / VAR  — EXECUTED orders tagged `h_…` / `v_…`, so both are
                       slices of Executed and never exceed it. A cancelled
                       hedge placed nothing in the market and is counted in
                       Failed, with every other cancellation

    Deduplication matches the trade value exactly: the COMPLETE rows are
    deduped as their own set (so Executed ties to the KPI), the non-COMPLETE
    rows as another.

    Only OPTION orders count — the same scope as the strikes section. A
    futures row is not index-options activity: one rejected
    BANKNIFTY28JUL26FUT from a NIFTY-algo account would otherwise invent an
    algo row in an index that algo never traded, and break the user-count
    reconciliation with the Trade Value tables (which see no executed order
    from it).

    The (algo, type) of an order is resolved EXACTLY as aggregate() resolves
    it for the trade value rows, so the two tables always agree. In
    particular a user with no MTM entry has neither — they collect in one
    blank "—" group instead of inheriting their server's algo, which would
    invent an Int block under an algo that is entirely Pos+Int."""
    allocations = allocations or {}
    type_map = type_map or {}
    type_index = _type_index(type_map)

    complete = [o for o in orders if o.status == "COMPLETE"]
    rest = [o for o in orders if o.status != "COMPLETE"]
    deduped = (dedup_orders(complete, "orders summary · executed")
               + dedup_orders(rest, "orders summary · not executed"))

    is_option = {}
    groups = {}
    for row in deduped:
        opt = is_option.get(row.symbol)
        if opt is None:
            opt = is_option[row.symbol] = parse_strike(row.symbol, bands) is not None
        if not opt:
            continue
        user_key = _user_key(row.user_id)
        entry = _pick_allocation(allocations, user_key, row.server)
        # The SERVER shown is the matched account's, not the order's. They can
        # disagree: a handful of orders arrive tagged with a server the account
        # does not live on (cross-panel bleed — every MStech panel shares one
        # server, so exports cross-fetch other panels' rows). Taking the algo
        # from the MTM and the server from the order paired VS23's algo 1 with
        # the name VS5, reading as "VS5 runs algo 1" when VS5 is wholly algo 8.
        # One source for both keeps the pairing truthful; server_mismatches()
        # reports the orders this hides.
        server_key = _server_key(entry["server"] if entry else row.server)
        algo = entry["algo"] if entry else ""
        user_type = (_resolve_type(type_map, type_index,
                                   _user_key(entry["user_id"]), server_key)
                     if entry else "")
        group = groups.setdefault((algo, user_type), {"users": set(), "servers": {}})
        group["users"].add(user_key)
        bucket = group["servers"].setdefault(
            server_key, {"users": set(), "orders": 0, "executed": 0,
                         "failed": 0, "pending": 0, "hedge": 0, "var": 0})
        bucket["users"].add(user_key)
        bucket["orders"] += 1
        # STATUS first, then the tag sub-divides only what executed — so Hedge
        # and VAR are slices of Executed, never of Total Orders. Counting the
        # tag across all statuses (the previous behaviour) reported hedge
        # orders that were cancelled as hedge activity: on 11-08 that read
        # 1,82,855 hedge against 3,09,760 executed.
        if row.status == "COMPLETE":
            bucket["executed"] += 1
            if row.category == "hedge":
                bucket["hedge"] += 1
            elif row.category == "var":
                bucket["var"] += 1
        elif row.status in PENDING_STATUSES:
            bucket["pending"] += 1     # still live — not a failure
        else:
            bucket["failed"] += 1

    out = []
    for (algo, user_type), group in groups.items():
        servers = [{"server": server, "users": len(b["users"]), "orders": b["orders"],
                    "executed": b["executed"], "failed": b["failed"],
                    "pending": b["pending"], "hedge": b["hedge"], "var": b["var"]}
                   for server, b in sorted(group["servers"].items())]
        out.append({
            "algo": algo,
            "user_type": user_type,
            # distinct across the algo's servers — a user on two servers counts once
            "users": len(group["users"]),
            # the ids themselves, so the blank-algo group can name the users
            # that fired orders but are in no User MTM
            "user_ids": sorted(group["users"]),
            "orders": sum(s["orders"] for s in servers),
            "executed": sum(s["executed"] for s in servers),
            "failed": sum(s["failed"] for s in servers),
            "pending": sum(s["pending"] for s in servers),
            "hedge": sum(s["hedge"] for s in servers),
            "var": sum(s["var"] for s in servers),
            "servers": servers,
        })
    # numeric algos first in numeric order, then text algos, blank (unmatched) last
    out.sort(key=lambda r: (
        (2, "") if r["algo"] == "" else
        (0, int(r["algo"])) if str(r["algo"]).isdigit() else (1, str(r["algo"])),
        r["user_type"],
    ))
    return out


def order_summary_totals(rows):
    """The all-algo total row for the orders summary. Users are distinct
    across algos only in so far as the rows already are — a user belongs to
    one (algo, type) group, so the counts add up."""
    return {
        "users": sum(r["users"] for r in rows),
        "orders": sum(r["orders"] for r in rows),
        "executed": sum(r["executed"] for r in rows),
        "failed": sum(r["failed"] for r in rows),
        "pending": sum(r["pending"] for r in rows),
        "hedge": sum(r["hedge"] for r in rows),
        "var": sum(r["var"] for r in rows),
    }


def format_order_row(row):
    # blank algo / type = a user with no MTM entry, shown as an explicit dash
    return [row["algo"] or "—", row["user_type"] or "—", row["users"], row["orders"],
            row["executed"], row["failed"], row["pending"], row["hedge"], row["var"]]


def add_orders_sheet(workbook, order_rows, suffix=""):
    """Append the "Orders" sheet: one row per (algo, type) with its servers
    listed underneath, then the grand total. `suffix` (e.g. " 2") names the
    second User MTM's sheet, matching the Segregation / summary sets."""
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    sheet = workbook.create_sheet("Orders" + suffix)
    bold = Font(bold=True)
    header = ORDER_SUMMARY_HEADER
    for col_idx, title in enumerate(header, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=title)
        cell.font = bold
        sheet.column_dimensions[get_column_letter(col_idx)].width = max(len(title) + 3, 12)

    r = 2
    for row in order_rows:
        for col_idx, value in enumerate(format_order_row(row), start=1):
            sheet.cell(row=r, column=col_idx, value=value).font = bold
        r += 1
        for s in row["servers"]:
            # server rows sit under their algo, indented in the Type column
            values = ["", f"  {s['server']}", s["users"], s["orders"],
                      s["executed"], s["failed"], s["pending"], s["hedge"], s["var"]]
            for col_idx, value in enumerate(values, start=1):
                sheet.cell(row=r, column=col_idx, value=value)
            r += 1

    totals = order_summary_totals(order_rows)
    for col_idx, value in enumerate(
            ["Total", "", totals["users"], totals["orders"], totals["executed"],
             totals["failed"], totals["pending"], totals["hedge"],
             totals["var"]], start=1):
        sheet.cell(row=r, column=col_idx, value=value).font = bold
    sheet.freeze_panes = "A2"


# Dot series on the intraday chart. STATUS decides first — an order that never
# executed placed nothing in the market — and the TAG then sub-divides only the
# orders that did complete:
#
#     complete = every executed lot            (= stoxxo + hedge + var)
#       stoxxo   executed, normal tag
#       hedge    executed, h_ tag
#       var      executed, v_ tag
#     failed   = every cancelled / rejected lot, WHATEVER the tag
#
# These OVERLAP by design: `complete` is the total and the next three are its
# parts, so the legend must not be summed. An earlier version checked the tag
# first, which put failed hedges into `hedge` — on 11-08 that inflated hedge to
# 15.6 lakh lots against a 5.6 lakh executed book, and left `failed` showing
# 1.1% of real failures because only normal-tagged failures reached it.
LOT_CATEGORIES = ("complete", "stoxxo", "hedge", "var", "failed")

# The parts of `complete`; used to assert the split adds back up.
LOT_EXECUTED_PARTS = ("stoxxo", "hedge", "var")


def _lot_categories(order):
    """Every series an order belongs to — a list, because `complete` overlaps
    its own parts. Pending orders are still live and belong to none of them,
    the same way the Orders Summary treats them as neither executed nor
    failed."""
    if order.status == "COMPLETE":
        if order.category == "hedge":
            return ["complete", "hedge"]
        if order.category == "var":
            return ["complete", "var"]
        return ["complete", "stoxxo"]
    if order.status in PENDING_STATUSES:
        return []
    return ["failed"]


def lots_timeline(orders, allocations=None, deduped=True):
    """{index: {"algos", "cats", "rows"}} — lots fired per minute, split by algo
    and by the colour category above.

    `rows` is [[minute, algo index, category index, lots], ...] with minute as
    "HH:MM", compact enough to embed and let the browser re-bucket to any
    timeframe and filter by algo without a round trip.

    Square-off orders are excluded: they close a position rather than place
    one, and on an expiry day they would otherwise dominate the last buckets.
    Orders whose file carried no usable time are skipped and counted."""
    allocations = allocations or {}
    per_index, skipped = {}, 0

    for row in orders:
        if not row.minute or row.category == "sqoff":
            skipped += not row.minute
            continue
        cats = _lot_categories(row)
        if not cats:
            continue                      # pending: live, in none of the series
        entry = _pick_allocation(allocations, _user_key(row.user_id), row.server)
        algo = str(entry["algo"]) if entry and entry["algo"] else "—"
        lot_size = _LOT_SIZE_BY_SEGMENT[row.segment](row.trade_date)
        lots = abs(row.qty) / lot_size
        bucket = per_index.setdefault(row.segment, {})
        for cat in cats:                  # `complete` AND the part it belongs to
            key = (row.minute, algo, cat)
            bucket[key] = bucket.get(key, Decimal("0")) + lots

    out = {}
    for index, buckets in per_index.items():
        algos = sorted({k[1] for k in buckets},
                       key=lambda a: (0, int(a)) if a.isdigit() else (1, a))
        algo_at = {a: i for i, a in enumerate(algos)}
        cat_at = {c: i for i, c in enumerate(LOT_CATEGORIES)}
        rows = [[minute, algo_at[algo], cat_at[cat], round(float(lots), 2)]
                for (minute, algo, cat), lots in sorted(buckets.items())]
        out[index] = {"algos": algos, "cats": list(LOT_CATEGORIES), "rows": rows}
    if skipped:
        logger.warning("lots timeline: %d order(s) had no readable Order Time "
                       "and are absent from the chart", skipped)
    return out


def volume_timeline(orders):
    """{index: [[minute, quantity], ...]} — traded QUANTITY per minute, our own.

    Quantity, not lots, so it is directly comparable with an exchange volume
    figure (exchanges report contracts). Completed orders only: volume means
    what transacted, so a cancelled order contributed nothing to the tape.
    Square-off is included — it is a real trade — unlike the lots panel, which
    excludes it because there the question is what was *placed*."""
    per_index = {}
    for row in orders:
        if row.status != "COMPLETE" or not row.minute:
            continue
        bucket = per_index.setdefault(row.segment, {})
        bucket[row.minute] = bucket.get(row.minute, Decimal("0")) + abs(row.qty)
    return {index: [[m, float(q)] for m, q in sorted(b.items())]
            for index, b in per_index.items()}


def strike_report(orders, allocations=None, bands=None):
    """The two strike summaries, computed over (deduped) orders:

      * by_algo_server — per (algo, server): how many DISTINCT contracts
        (segment, strike, CE/PE) were traded. The algo comes from the
        same MTM allocation matching as the trade value rows; orders whose
        user has no MTM entry land in a blank-algo bucket.
      * per_strike — per contract: total lots (sum |qty| / lot size, the same
        date-based lot math as the trade value rows) and order count.

    A contract has no expiry axis — one orderbook holds a single expiry per
    index, entered in the dashboard and applied as a display label there.
    Orders whose symbol doesn't parse as an option contract are skipped."""
    allocations = allocations or {}
    # Server -> algo fallback for users with no MTM entry: a server's algo is
    # the algo of its MTM accounts (the most common one, should they ever
    # mix), so an unmatched user doesn't spawn a duplicate blank-algo row
    # for a server that already sits under an algo.
    server_algo_votes = {}
    for entries in allocations.values():
        for entry in entries:
            if entry["algo"]:
                server_algo_votes.setdefault(entry["server"], Counter())[entry["algo"]] += 1
    server_algo = {server: votes.most_common(1)[0][0]
                   for server, votes in server_algo_votes.items()}

    parse_cache = {}
    for row in orders:
        if row.symbol not in parse_cache:
            parse_cache[row.symbol] = parse_strike(row.symbol, bands)

    by_algo_server = {}
    per_strike = {}
    for row in orders:
        contract = parse_cache[row.symbol]
        if contract is None:
            continue
        entry = _pick_allocation(allocations, _user_key(row.user_id), row.server)
        server_key = _server_key(row.server)
        algo = entry["algo"] if entry else server_algo.get(server_key, "")
        by_algo_server.setdefault((algo, server_key), set()).add(contract)
        lot_size = _LOT_SIZE_BY_SEGMENT[row.segment](row.trade_date)
        group = per_strike.setdefault(contract, {"lots": Decimal("0"), "order_count": 0,
                                                 "breakdown": {}})
        lots = abs(row.qty) / lot_size
        group["lots"] += lots
        group["order_count"] += 1
        # per (algo, trade kind) split — feeds the chain's hedge/var/sqoff/algo filters
        bkey = (algo, row.category)
        group["breakdown"][bkey] = group["breakdown"].get(bkey, Decimal("0")) + lots

    # each row keeps its contract set so the count can be re-filtered by
    # expiry / index in the dashboard and DOR.html
    algo_rows = [
        {"algo": algo, "server": server, "strike_count": len(contracts),
         "contracts": sorted(contracts)}
        for (algo, server), contracts in by_algo_server.items()
    ]
    # numeric algos first in numeric order, then text algos, blank (unmatched) last
    algo_rows.sort(key=lambda r: (
        (2, "") if r["algo"] == "" else
        (0, int(r["algo"])) if str(r["algo"]).isdigit() else (1, str(r["algo"])),
        r["server"],
    ))
    strike_rows = [
        {"segment": c[0], "strike": c[1], "opt_type": c[2],
         "lots": g["lots"], "order_count": g["order_count"],
         "breakdown": g["breakdown"]}
        for c, g in per_strike.items()
    ]
    strike_rows.sort(key=lambda r: (r["segment"], r["strike"], r["opt_type"]))
    return {"by_algo_server": algo_rows, "per_strike": strike_rows}


def strike_chain(per_strike):
    """Pivot the per-strike rows into an option-chain view:
    {segment: [(ce_lots, strike, pe_lots), ...]} — one row per strike price,
    sorted by strike, CE and PE lots side by side (0 when only one side
    traded). Keys come out sorted by segment."""
    by_key = {}
    for row in per_strike:
        sides = by_key.setdefault(row["segment"], {}) \
                      .setdefault(row["strike"], {"CE": Decimal("0"), "PE": Decimal("0")})
        sides[row["opt_type"]] += row["lots"]
    return {
        segment: [(sides[s]["CE"], s, sides[s]["PE"]) for s in sorted(sides)]
        for segment, sides in sorted(by_key.items())
    }


def add_strikes_sheet(workbook, strikes, expiry_map=None):
    """Append the "Strikes" sheet: the per-(algo, server) distinct strike
    counts in columns A-C, and one option-chain block (CE | Strike | PE) per
    segment stacked from column E. `expiry_map` ({index: expiry text entered
    in the dashboard}) captions each block. Lots are whole numbers."""
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    sheet = workbook.create_sheet("Strikes")
    bold = Font(bold=True)

    for col_idx, title in enumerate(STRIKE_ALGO_HEADER, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=title)
        cell.font = bold
        sheet.column_dimensions[get_column_letter(col_idx)].width = max(len(title) + 4, 12)
    for r, row in enumerate(strikes["by_algo_server"], start=2):
        sheet.cell(row=r, column=1, value=row["algo"])
        sheet.cell(row=r, column=2, value=row["server"])
        sheet.cell(row=r, column=3, value=row["strike_count"])
    # a per-(algo, server) count can share strikes with another server, so the
    # overall figure is the distinct count, not the column sum
    total_row = len(strikes["by_algo_server"]) + 2
    sheet.cell(row=total_row, column=1, value="Total (distinct)").font = bold
    sheet.cell(row=total_row, column=3, value=len(strikes["per_strike"])).font = bold

    offset = len(STRIKE_ALGO_HEADER) + 2  # leave column D blank between the tables
    for j in range(3):
        sheet.column_dimensions[get_column_letter(offset + j)].width = 12
    r = 1
    for segment, rows in strike_chain(strikes["per_strike"]).items():
        label = (expiry_map or {}).get(segment, "")
        sheet.merge_cells(start_row=r, start_column=offset, end_row=r, end_column=offset + 2)
        title = sheet.cell(row=r, column=offset,
                           value=f"{segment} {label}".rstrip())
        title.font = bold
        r += 1
        for j, header in enumerate(STRIKE_CHAIN_HEADER):
            sheet.cell(row=r, column=offset + j, value=header).font = bold
        r += 1
        ce_total = pe_total = Decimal("0")
        for ce, strike, pe in rows:
            sheet.cell(row=r, column=offset, value=int(round(ce))).number_format = INDIAN_XLSX_FMT
            sheet.cell(row=r, column=offset + 1, value=strike)
            sheet.cell(row=r, column=offset + 2, value=int(round(pe))).number_format = INDIAN_XLSX_FMT
            ce_total += ce
            pe_total += pe
            r += 1
        for j, value in enumerate((int(round(ce_total)), "Total", int(round(pe_total)))):
            cell = sheet.cell(row=r, column=offset + j, value=value)
            cell.font = bold
            if j != 1:
                cell.number_format = INDIAN_XLSX_FMT
        r += 2  # blank spacer row between chains
    sheet.freeze_panes = "A2"


def report_csv_text(rows):
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(REPORT_HEADER)
    for row in rows:
        writer.writerow(format_row(row))
    return buffer.getvalue()


def _excel_number(value, places=2):
    if value is None:
        return None
    return round(float(value), places)


def _excel_report_row(row):
    # User id stays a string cell so Excel keeps leading zeros ("04101961").
    return [
        row["trade_date"].strftime("%d-%m-%Y") if row["trade_date"] else "",
        row["server"],
        row["user_id"],
        _excel_number(row["allocation"], places=0),
        row["algo"],
        row.get("user_type", ""),
        row["segment"],
        row["order_count"],
        _excel_number(row["quantity"], places=0),
        _excel_number(row["lots"], places=0),
        _excel_number(row.get("normalise")),
        _excel_number(row.get("lots_per_cr")),
        _excel_number(row["trade_value"]),
        row.get("outlier", ""),
    ]


def add_report_sheets(workbook, rows, std_multiplier=None, summary_groups=None):
    """Append the "tradevalue" and "summary" sheets to an existing openpyxl
    workbook (used to combine this report with other reports in one file).
    `summary_groups` ([(sheet title, rows), ...]) writes one summary sheet
    per group — used when a secondary User MTM splits the outlier bands."""
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    sheet = workbook.create_sheet("tradevalue")
    sheet.append(REPORT_HEADER)
    for row in rows:
        sheet.append(_excel_report_row(row))

    bold = Font(bold=True)
    sized = [(sheet, REPORT_HEADER)]
    for title, group_rows in summary_groups or [("summary", rows)]:
        summary_sheet = workbook.create_sheet(title)
        summary_sheet.append(SUMMARY_HEADER)
        for srow in algo_summary(group_rows, std_multiplier):
            has_stats = srow["median_lots_per_cr"] is not None
            summary_sheet.append([
                srow["algo"],
                srow["user_type"],
                srow["users"],
                _excel_number(srow["median_lots_per_cr"]),
                (f"{_fmt_decimal(srow['band_low'])}–{_fmt_decimal(srow['band_high'])}"
                 if has_stats else ""),
                srow["below"],
                srow["in_range"],
                srow["above"],
            ])
        sized.append((summary_sheet, SUMMARY_HEADER))

    for ws, header in sized:
        for col_idx, title in enumerate(header, start=1):
            ws.cell(row=1, column=col_idx).font = bold
            ws.column_dimensions[get_column_letter(col_idx)].width = max(len(title) + 4, 12)
        ws.freeze_panes = "A2"


def write_report_excel(rows, target, std_multiplier=None, strikes=None):
    """Write the report workbook: sheet "tradevalue" holds the full report,
    sheet "summary" the per-algo outlier table, plus the "Strikes" sheet when
    strike data is given. `target` may be a filesystem path or a writable
    buffer."""
    from openpyxl import Workbook

    workbook = Workbook()
    workbook.remove(workbook.active)
    add_report_sheets(workbook, rows, std_multiplier)
    if strikes:
        add_strikes_sheet(workbook, strikes)
    workbook.save(target)


def report_excel_bytes(rows, std_multiplier=None, strikes=None):
    buffer = io.BytesIO()
    write_report_excel(rows, buffer, std_multiplier, strikes)
    return buffer.getvalue()


def write_report(rows, output_path, std_multiplier=None, strikes=None):
    if str(output_path).lower().endswith(".csv"):
        with open(output_path, "w", newline="", encoding="utf-8") as fh:
            fh.write(report_csv_text(rows))
    else:
        write_report_excel(rows, output_path, std_multiplier, strikes)


def report_totals(rows):
    return {
        "users": len({r["user_id"] for r in rows if r["user_id"]}),
        "orders": sum(r["order_count"] for r in rows),
        "lots": sum((r["lots"] for r in rows), Decimal("0")),
        "trade_value": sum((r["trade_value"] for r in rows), Decimal("0")),
    }


def build_report(orderbook_source, summary_source, orderbook_name=None, summary_name=None):
    """Full pipeline: read both inputs, dedup, aggregate. Returns report rows."""
    orders = read_orderbook(orderbook_source, orderbook_name)
    allocations = read_allocations(summary_source, summary_name) if summary_source is not None else {}
    return aggregate(dedup_orders(orders), allocations)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _default_files(pattern):
    return [p for p in sorted(SCRIPT_DIR.glob(pattern)) if p.suffix.lower() in TABLE_SUFFIXES]


def main(argv=None):
    parser = argparse.ArgumentParser(description="Generate the trade value report from orderbook + user MTM files.")
    parser.add_argument("inputs", nargs="*", type=Path,
                        help=f"orderbook CSV/Excel file(s); default: {ORDERBOOK_PATTERN} in the script folder")
    parser.add_argument("-s", "--summary", type=Path,
                        help=f"compiled user MTM summary (CSV/Excel) for the Allocation column; "
                             f"default: {SUMMARY_PATTERN} in the script folder")
    parser.add_argument("-o", "--output", type=Path, default=DEFAULT_OUTPUT,
                        help=f"output path, .xlsx or .csv (default: {DEFAULT_OUTPUT.name})")
    parser.add_argument("-d", "--deviation", type=float, default=1.0,
                        help="outlier threshold in standard deviations of the algo's lot pct (default 1)")
    args = parser.parse_args(argv)

    if args.deviation <= 0:
        parser.error("--deviation must be greater than 0")
    std_multiplier = Decimal(str(args.deviation))

    inputs = args.inputs or _default_files(ORDERBOOK_PATTERN)
    if not inputs:
        parser.error(f"no orderbook files found ({ORDERBOOK_PATTERN})")
    missing = [p for p in inputs if not p.is_file()]
    if missing:
        parser.error("input file not found: " + ", ".join(str(p) for p in missing))

    summary = args.summary
    if summary is None:
        candidates = _default_files(SUMMARY_PATTERN)
        summary = candidates[0] if candidates else None
        if summary is None:
            print(f"WARNING: no summary file found ({SUMMARY_PATTERN}); Allocation column will be empty")
    elif not summary.is_file():
        parser.error(f"summary file not found: {summary}")

    orders = []
    for path in inputs:
        file_orders = read_orderbook(path)
        print(f"{path.name}: {len(file_orders)} completed NIFTY/BANKNIFTY/SENSEX orders")
        orders.extend(file_orders)

    allocations = {}
    if summary is not None:
        allocations = read_allocations(summary)
        print(f"{summary.name}: allocations for {len(allocations)} users")

    deduped = dedup_orders(orders)
    dropped = len(orders) - len(deduped)
    if dropped:
        print(f"dropped {dropped} duplicate orders")

    report_rows = aggregate(deduped, allocations, std_multiplier)
    strikes = strike_report(deduped, allocations)
    write_report(report_rows, args.output, std_multiplier, strikes)

    totals = report_totals(report_rows)
    matched = len({r["user_id"] for r in report_rows if r["allocation"] is not None})
    print(f"wrote {len(report_rows)} rows to {args.output}")
    print(f"allocation matched for {matched}/{totals['users']} users")
    print(f"totals: {totals['users']} users | lots {format_indian(totals['lots'])} "
          f"| trade value {format_indian(totals['trade_value'], places=2)}")
    print(f"strikes: {len(strikes['per_strike'])} distinct contracts")


if __name__ == "__main__":
    main()


def server_mismatches(orders, allocations):
    """{(user, order server, MTM server, algo): order count} where an order's
    server is not the server its account lives on.

    Small and worth seeing: these are the orders whose server is rewritten for
    display by order_summary. On 20-08-2026 there were 98 of 475,592, all one
    account family, and they are the reason a VS5 row can carry a non-A8 algo."""
    out = {}
    for row in orders:
        key = _user_key(row.user_id)
        entry = _pick_allocation(allocations, key, row.server)
        if entry is None:
            continue
        order_server = _server_key(row.server)
        if entry["server"] and order_server and entry["server"] != order_server:
            k = (key, order_server, entry["server"], entry["algo"])
            out[k] = out.get(k, 0) + 1
    return out


def drop_unmatched(orders, allocations):
    """(kept orders, {(user key, server): count}) — orders whose account is in
    neither the User MTM nor aliases.json are removed from every calculation.

    Call AFTER add_user_aliases, so an id that only resolves through an alias
    is kept. An unmatched id carries no algo, no allocation and no type, so it
    can only ever land in the "-" row; counting its orders in the totals while
    attributing them to nothing makes the report add up to more than it can
    explain. Dropping is the desk's stated rule.

    This CHANGES the headline totals, so the caller is expected to report what
    went — silently shrinking an order count is worse than the "-" row it
    replaces."""
    kept, dropped = [], {}
    for row in orders:
        key = _user_key(row.user_id)
        if _pick_allocation(allocations, key, row.server) is None:
            k = (key, _server_key(row.server))
            dropped[k] = dropped.get(k, 0) + 1
            continue
        kept.append(row)
    if dropped:
        logger.warning("dropped %d order(s) from %d account(s) absent from the "
                       "User MTM and aliases.json: %s", sum(dropped.values()),
                       len(dropped),
                       ", ".join(f"{u} on {s}" for u, s in sorted(dropped)[:6]))
    return kept, dropped
